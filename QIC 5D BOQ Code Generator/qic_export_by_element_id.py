#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
BIM DATA EXPORT SYSTEM  —  v10.0
Reads: IFC (.ifc), Navisworks Cache (.nwc), Revit (.rvt)
Exports: Excel (.xlsx) matching exact WPH-BUJV reference format
Author: Ahmed Khalaf  —  UCC BIM Manager
"""

import os, re, sys, io, struct, json, time, uuid
from datetime import datetime
from collections import defaultdict, OrderedDict
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

RUNNING_IN_REVIT = False
try:
    import clr
    import System

    clr.AddReference('RevitAPI')
    from Autodesk.Revit.DB import *

    clr.AddReference('RevitServices')
    from RevitServices.Persistence import DocumentManager

    RUNNING_IN_REVIT = True
except ImportError:
    pass

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                 GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("Installing openpyxl…")
    import subprocess
    installed = False
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        installed = True
    except Exception:
        pass

    if not installed:
        try:
            subprocess.check_call(["uv", "pip", "install", "openpyxl"])
            installed = True
        except Exception:
            pass

    if not installed:
        try:
            subprocess.check_call([sys.executable, "-m", "ensurepip", "--default-pip"])
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
            installed = True
        except Exception:
            pass

    if not installed:
        raise ImportError("Could not install openpyxl automatically. Please run: uv pip install openpyxl")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS  —  colours (openpyxl = no #,  tkinter = with #)
# ─────────────────────────────────────────────────────────────────────────────
# openpyxl hex colours (no hash)
C_BLUE_DARK = "1a73e8"
C_BLUE_LIGHT = "e8f0fe"
C_WHITE = "FFFFFF"
C_ORANGE = "e67700"
C_GREY_LABEL = "555555"
C_GREY_VALUE = "444444"
C_AUTHOR = "1a73e8"
C_HEADER_FG = "FFFFFF"
C_ROW_ALT = "e8f0fe"
C_ROW_PLAIN = "FFFFFF"

# tkinter colours (with hash — ttk/tk require this)
TK_BLUE = "#1a73e8"
TK_BLUE_LT = "#e8f0fe"
TK_BLUE_DK = "#1557b0"
TK_WHITE = "#ffffff"
TK_BG = "#f4f6fb"
TK_PANEL = "#ffffff"
TK_ORANGE = "#e67700"
TK_GREY = "#555555"
TK_GREY_LT = "#888888"
TK_GREEN = "#1e8e3e"
TK_GREEN_HV = "#166d35"
TK_SLATE = "#455a64"
TK_SLATE_HV = "#37474f"
TK_PURPLE = "#5c6bc0"
TK_PURPLE_HV = "#3949ab"
TK_ROW_ALT = "#e8f0fe"
TK_ROW_PLAIN = "#ffffff"
TK_BORDER = "#dadce0"
TK_SUCCESS = "#34a853"
TK_ERROR = "#ea4335"
TK_WARNING = "#fbbc04"

AUTHOR = "Ahmed Khalaf  —  UCC BIM Manager"

# ─────────────────────────────────────────────────────────────────────────────
# QIC SHARED PARAMETER CATALOGUE  (37 standard params)
# ─────────────────────────────────────────────────────────────────────────────
QIC_SHARED_PARAMS = [
    {"guid": "a1b2c3d4-0001-0000-0000-000000000001", "name": "QIC_IDENTITY_CODE", "type": "Text",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "Unique QIC identity code", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000002", "name": "QIC_ASSET_TAG", "type": "Text", "group": "Identity Data",
     "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common", "desc": "Physical asset tag / barcode",
     "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000003", "name": "QIC_CLASSIFICATION_CODE", "type": "Text",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "OmniClass / Uniclass code", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000004", "name": "QIC_LOD", "type": "Integer", "group": "Construction",
     "code": "PG_CONSTRUCTION", "mod": True, "vis": True, "disc": "Common", "desc": "Level of Detail 100-500",
     "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000005", "name": "QIC_LOI", "type": "Integer", "group": "Construction",
     "code": "PG_CONSTRUCTION", "mod": True, "vis": True, "disc": "Common", "desc": "Level of Information 1-5",
     "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000006", "name": "QIC_STATUS", "type": "Text", "group": "Identity Data",
     "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common", "desc": "Element status",
     "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000007", "name": "QIC_PHASE", "type": "Text", "group": "Phasing",
     "code": "PG_PHASING", "mod": True, "vis": True, "disc": "Common", "desc": "Construction phase",
     "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000008", "name": "QIC_ZONE", "type": "Text", "group": "Identity Data",
     "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common", "desc": "Zone code",
     "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000009", "name": "QIC_SYSTEM", "type": "Text", "group": "Mechanical",
     "code": "PG_MECHANICAL", "mod": True, "vis": True, "disc": "HVAC", "desc": "System classification",
     "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000010", "name": "QIC_MANUFACTURER", "type": "Text",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "Manufacturer name", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000011", "name": "QIC_MODEL_NUMBER", "type": "Text",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "Model / catalogue number", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000012", "name": "QIC_MATERIAL_GRADE", "type": "Text", "group": "Materials",
     "code": "PG_MATERIALS", "mod": True, "vis": True, "disc": "Structural", "desc": "Material specification grade",
     "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000013", "name": "QIC_FIRE_RATING", "type": "Text",
     "group": "Fire Protection", "code": "PG_FIRE_PROTECTION", "mod": True, "vis": True, "disc": "Common",
     "desc": "Fire resistance rating", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000014", "name": "QIC_ACOUSTIC_RATING", "type": "Number",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "Sound transmission class dB", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000015", "name": "QIC_THERMAL_VALUE", "type": "Number", "group": "Thermal",
     "code": "PG_THERMAL", "mod": True, "vis": True, "disc": "Energy", "desc": "U-value / R-value",
     "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000016", "name": "QIC_COST_CODE", "type": "Text", "group": "Construction",
     "code": "PG_CONSTRUCTION", "mod": True, "vis": True, "disc": "Common", "desc": "Cost centre / BOQ code",
     "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000017", "name": "QIC_WARRANTY_PERIOD", "type": "Integer",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "Warranty period months", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000018", "name": "QIC_MAINTENANCE_CYCLE", "type": "Text",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "Maintenance frequency", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000019", "name": "QIC_INSTALL_DATE", "type": "Text",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "Installation date YYYY-MM-DD", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000020", "name": "QIC_SPATIAL_LOCATION", "type": "Text",
     "group": "Constraints", "code": "PG_CONSTRAINTS", "mod": True, "vis": True, "disc": "Common",
     "desc": "Spatial / room location ref", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000021", "name": "QIC_DRAWING_REF", "type": "Text",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "Originating drawing number", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000022", "name": "QIC_SPEC_SECTION", "type": "Text",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "Specification section ref", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000023", "name": "QIC_DOCUMENT_REF", "type": "Text",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "Related document reference", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000024", "name": "QIC_SUBMITTAL_REF", "type": "Text",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "Submittal / RFI reference", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000025", "name": "QIC_INSPECTION_STATUS", "type": "Text",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "QA inspection status", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000026", "name": "QIC_HANDOVER_STATUS", "type": "Text",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "O&M handover status", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000027", "name": "QIC_CONTRACTOR", "type": "Text", "group": "Identity Data",
     "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common", "desc": "Installing contractor",
     "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000028", "name": "QIC_SUBCONTRACTOR", "type": "Text",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "Sub-contractor name", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000029", "name": "QIC_SUPPLIER", "type": "Text", "group": "Identity Data",
     "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common", "desc": "Material supplier name",
     "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000030", "name": "QIC_COUNTRY_OF_ORIGIN", "type": "Text",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "Country of manufacture", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000031", "name": "QIC_LEED_CREDIT", "type": "Text",
     "group": "Green Building", "code": "PG_GREEN_BUILDING", "mod": True, "vis": True, "disc": "Energy",
     "desc": "LEED credit contribution", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000032", "name": "QIC_RECYCLED_CONTENT", "type": "Number",
     "group": "Green Building", "code": "PG_GREEN_BUILDING", "mod": True, "vis": True, "disc": "Energy",
     "desc": "Recycled content percentage", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000033", "name": "QIC_STRUCTURAL_GRADE", "type": "Text",
     "group": "Structural", "code": "PG_STRUCTURAL", "mod": True, "vis": True, "disc": "Structural",
     "desc": "Structural design grade", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000034", "name": "QIC_LOAD_CAPACITY", "type": "Number",
     "group": "Structural", "code": "PG_STRUCTURAL", "mod": True, "vis": True, "disc": "Structural",
     "desc": "Design load capacity kN", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000035", "name": "QIC_SERVICE_LIFE", "type": "Integer",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "Design service life years", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000036", "name": "QIC_NATIONAL_ANNEX", "type": "Text",
     "group": "Identity Data", "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common",
     "desc": "National annex reference", "source": "QIC BIM Standard v3"},
    {"guid": "a1b2c3d4-0001-0000-0000-000000000037", "name": "QIC_NOTES", "type": "Text", "group": "Identity Data",
     "code": "PG_IDENTITY_DATA", "mod": True, "vis": True, "disc": "Common", "desc": "Free-text notes / remarks",
     "source": "QIC BIM Standard v3"},
]

# ─────────────────────────────────────────────────────────────────────────────
# CATEGORY COLUMN DEFINITIONS
# ─────────────────────────────────────────────────────────────────────────────
# Base columns present on every category sheet (IFC-aligned)
BASE_COLS = [
    "Element ID", "IFC Class", "Category", "Name", "Description",
    "Object Type", "Mark / Tag", "Predefined Type", "IFC GlobalId", "Level",
]

# Extra columns per category (matching exact reference headers)
CATEGORY_EXTRA_COLS = {
    "Walls": ["[Pset_ProductRequirements] Category", "Mark", "External",
              "[Pset_WallCommon] ExtendToStructure", "Structural"],
    "Floors": ["[Pset_ProductRequirements] Category", "Mark", "Structural",
               "External", "[Pset_SlabCommon] PitchAngle"],
    "Stairs": ["[Pset_ProductRequirements] Category", "Mark", "External",
               "Structural", "Multistory Top Level", "Base Level",
               "Base Offset", "Top Level", "Top Offset",
               "Actual Run Width", "Actual Tread Depth",
               "Maximum Riser Height", "Minimum Tread Depth",
               "Minimum Run Width", "Nosing Length",
               "Number of Risers", "Actual Number of Risers",
               "Desired Number of Risers", "Maximum Number of Risers",
               "Number of Treads", "Actual Tread Depth (read-only)",
               "Riser Thickness", "Tread Thickness",
               "Right Lateral Offset", "Left Lateral Offset"],
    "Structural Columns": ["[Pset_ProductRequirements] Category", "Mark", "External",
                           "Structural", "[Pset_ColumnCommon] LoadBearing"],
    "Structural Framing": ["[Pset_ProductRequirements] Category", "Mark", "External",
                           "Structural", "[Pset_BeamCommon] LoadBearing",
                           "[Pset_BeamCommon] IsExternal"],
    "Topography": ["[Pset_ProductRequirements] Category", "Mark",
                   "Structural", "External", "[Pset_SiteCommon] BuildableArea",
                   "[Pset_SiteCommon] TotalArea"],
    "Levels": ["[Pset_ProductRequirements] Category", "Mark",
               "Elevation", "Story Above", "Computation Height",
               "Floor Type", "Building Story", "Structural"],
    "Doors": ["[Pset_ProductRequirements] Category", "Mark", "External",
              "[Pset_DoorCommon] IsExternal", "Fire Rating",
              "Acoustic Rating", "Width", "Height"],
    "Windows": ["[Pset_ProductRequirements] Category", "Mark", "External",
                "[Pset_WindowCommon] IsExternal", "Fire Rating",
                "Width", "Height"],
    "Ceilings": ["[Pset_ProductRequirements] Category", "Mark", "External",
                 "Structural", "[Pset_CoveringCommon] Thickness"],
    "Roofs": ["[Pset_ProductRequirements] Category", "Mark", "External",
              "Structural", "[Pset_RoofCommon] IsExternal"],
    "Ramps": ["[Pset_ProductRequirements] Category", "Mark", "External",
              "Structural", "[Pset_RampCommon] IsExternal"],
    "Furniture": ["[Pset_ProductRequirements] Category", "Mark"],
    "Mechanical Equipment": ["[Pset_ProductRequirements] Category", "Mark",
                             "[Pset_ManufacturerTypeInformation] Manufacturer"],
    "Plumbing Fixtures": ["[Pset_ProductRequirements] Category", "Mark",
                          "[Pset_SanitaryTerminalTypeToiletPan] ToiletType"],
    "Electrical Fixtures": ["[Pset_ProductRequirements] Category", "Mark"],
    "Lighting Fixtures": ["[Pset_ProductRequirements] Category", "Mark",
                          "[Pset_LightFixtureTypeCommon] LightFixtureType"],
    "Pipes": ["[Pset_ProductRequirements] Category", "Mark",
              "[Pset_PipeSegmentTypeCommon] NominalDiameter"],
    "Ducts": ["[Pset_ProductRequirements] Category", "Mark",
              "[Pset_DuctSegmentTypeCommon] NominalWidth"],
    "Spaces": ["[Pset_ProductRequirements] Category", "Mark",
               "[Pset_SpaceCommon] GrossFloorArea",
               "[Pset_SpaceCommon] NetFloorArea"],
    "Zones": ["[Pset_ProductRequirements] Category", "Mark"],
    "Generic Models": ["[Pset_ProductRequirements] Category", "Mark"],
    "Mass": ["[Pset_ProductRequirements] Category", "Mark"],
    "Site": ["[Pset_ProductRequirements] Category", "Mark"],
}

# IFC class → Category mapping
IFC_CLASS_MAP = {
    "IFCWALL": "Walls", "IFCWALLSTANDARDCASE": "Walls",
    "IFCSLAB": "Floors", "IFCFOOTING": "Floors",
    "IFCSTAIR": "Stairs", "IFCSTAIRCASEFLIGHT": "Stairs", "IFCSTAIRFLIGHT": "Stairs",
    "IFCCOLUMN": "Structural Columns",
    "IFCBEAM": "Structural Framing", "IFCMEMBER": "Structural Framing",
    "IFCSITE": "Topography",
    "IFCBUILDING": "Site",
    "IFCBUILDINGSTOREY": "Levels",
    "IFCDOOR": "Doors",
    "IFCWINDOW": "Windows",
    "IFCCOVERING": "Ceilings",
    "IFCROOF": "Roofs",
    "IFCRAMP": "Ramps", "IFCRAMPFLIGHT": "Ramps",
    "IFCFURNITURE": "Furniture", "IFCFURNISHINGELEMENT": "Furniture",
    "IFCFLOWSEGMENT": "Pipes",
    "IFCPIPESEGMENT": "Pipes",
    "IFCDUCTSEGMENT": "Ducts",
    "IFCFLOWFITTING": "Mechanical Equipment",
    "IFCENERGYCONVERSIONDEVICE": "Mechanical Equipment",
    "IFCFLOWMOVINGDEVICE": "Mechanical Equipment",
    "IFCFLOWTERMINAL": "Plumbing Fixtures",
    "IFCELECHEATINGELEMENT": "Electrical Fixtures",
    "IFCLIGHTFIXTURE": "Lighting Fixtures",
    "IFCSPACE": "Spaces",
    "IFCZONE": "Zones",
    "IFCBUILDINGELEMENTPROXY": "Generic Models",
    "IFCVIRTUALELEMENT": "Generic Models",
}


# ─────────────────────────────────────────────────────────────────────────────
# IFC PARSER  —  pure-Python STEP Physical File reader
# ─────────────────────────────────────────────────────────────────────────────
class IFCParser:
    """Parse IFC STEP Physical File without any external dependencies."""

    def __init__(self):
        self.entities = {}
        self.project_info = {}
        self.categories = {}
        self.shared_param_catalogue = []
        self.source = "IFC"
        self._pset_filter = None  # None → use DEFAULT_PSET_FILTER

    def parse(self, path: str):
        self.source = os.path.basename(path)
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            raw = f.read()

        # -- header project info --
        self._parse_header(raw)

        # -- data section entities --
        data_match = re.search(r"DATA;\s*(.*?)\s*ENDSEC;", raw, re.DOTALL | re.IGNORECASE)
        if not data_match:
            raise ValueError("No DATA section found in IFC file")

        line_re = re.compile(r"#(\d+)\s*=\s*([A-Z][A-Z0-9_]*)\s*\(([^;]*?)\)\s*;",
                             re.DOTALL)
        for m in line_re.finditer(data_match.group(1)):
            eid = int(m.group(1))
            cls = m.group(2).upper()
            raw_attrs = m.group(3)
            attrs = self._split_attrs(raw_attrs)
            self.entities[eid] = (cls, attrs)

        # -- resolve categories (pset_filter=None → use DEFAULT_PSET_FILTER) --
        self._build_categories(pset_filter=self._pset_filter)
        return self

    def set_pset_filter(self, pset_filter):
        """
        Configure which Psets to load before calling parse().
        pset_filter : set of strings  — pset name prefixes to include
                      None            — use DEFAULT_PSET_FILTER
                      set()           — include ALL psets (no filter)
        """
        self._pset_filter = pset_filter
        return self  # chainable: IFCParser().set_pset_filter({...}).parse(path)

    # ── header ────────────────────────────────────────────────────────────────
    def _parse_header(self, raw: str):
        hdr = re.search(r"HEADER;\s*(.*?)\s*ENDSEC;", raw, re.DOTALL | re.IGNORECASE)
        if not hdr:
            return
        h = hdr.group(1)
        fn = re.search(r"FILE_NAME\s*\(([^)]+)\)", h)
        fd = re.search(r"FILE_DESCRIPTION\s*\(([^)]+)\)", h)
        fsc = re.search(r"FILE_SCHEMA\s*\(\s*\(\s*'([^']+)'", h)

        if fn:
            parts = self._split_attrs(fn.group(1))
            self.project_info["file_name"] = self._unquote(parts[0]) if parts else ""
            self.project_info["time_stamp"] = self._unquote(parts[1]) if len(parts) > 1 else ""
            self.project_info["author"] = self._unquote(parts[2]) if len(parts) > 2 else ""
            self.project_info["organization"] = self._unquote(parts[3]) if len(parts) > 3 else ""
        if fsc:
            self.project_info["schema"] = fsc.group(1)

    # ── attribute splitter (handles nested parens & quoted strings) ────────────
    @staticmethod
    def _split_attrs(s: str):
        attrs, cur, depth, in_q = [], [], 0, False
        i = 0
        while i < len(s):
            c = s[i]
            if c == "'" and not in_q:
                in_q = True;
                cur.append(c)
            elif c == "'" and in_q:
                in_q = False;
                cur.append(c)
            elif not in_q and c == "(":
                depth += 1;
                cur.append(c)
            elif not in_q and c == ")":
                depth -= 1;
                cur.append(c)
            elif not in_q and c == "," and depth == 0:
                attrs.append("".join(cur).strip())
                cur = []
            else:
                cur.append(c)
            i += 1
        if cur:
            attrs.append("".join(cur).strip())
        return attrs

    @staticmethod
    def _unquote(s: str) -> str:
        s = s.strip()
        if s.startswith("'") and s.endswith("'"):
            return s[1:-1].replace("''", "'")
        return s

    def _ref(self, s: str):
        """Dereference a #N pointer, return (class, attrs) or None."""
        s = s.strip()
        if s.startswith("#"):
            try:
                return self.entities.get(int(s[1:]))
            except ValueError:
                return None
        return None

    def _attr(self, eid: int, idx: int) -> str:
        ent = self.entities.get(eid)
        if ent and len(ent[1]) > idx:
            return ent[1][idx]
        return ""

    # ── Pset filter configuration ──────────────────────────────────────────────
    # Only these Pset prefixes are loaded by default.
    # Pass pset_filter=None to IFCParser.parse() to load ALL psets.
    # Pass a set of strings to restrict — e.g. {"Pset_WallCommon", "Qto_WallBaseQuantities"}
    DEFAULT_PSET_FILTER: set = {
        "Pset_WallCommon",
        "Pset_SlabCommon",
        "Pset_ColumnCommon",
        "Pset_BeamCommon",
        "Pset_StairCommon",
        "Pset_RoofCommon",
        "Pset_DoorCommon",
        "Pset_WindowCommon",
        "Pset_CoveringCommon",
        "Pset_RampCommon",
        "Pset_SiteCommon",
        "Pset_SpaceCommon",
        "Pset_ProductRequirements",
        "Pset_ManufacturerTypeInformation",
        "Qto_WallBaseQuantities",
        "Qto_SlabBaseQuantities",
        "Qto_ColumnBaseQuantities",
        "Qto_BeamBaseQuantities",
    }

    # ── build categories — FULL DATA EXTRACTION ───────────────────────────────
    def _build_categories(self, pset_filter: set = None):
        """
        Extract EVERY data item assigned to every element:
          • All Psets / Qto sets  (IFCPROPERTYSET, IFCELEMENTQUANTITY)
          • All property types    (single-value, bounded, enum, list, table,
                                   complex, profile-def quantities)
          • Material layers       (IFCMATERIALLAYERSET / LAYERSETUSAGE)
          • Material constituents (IFCMATERIALCONSTITUENTSET — IFC4)
          • Type properties       (IFCRELDEFINESBYTYPE → type element psets)
          • Classification refs   (IFCRELASSOCIATESCLASSIFICATION)
          • Document refs         (IFCRELASSOCIATESDOCUMENT)
          • Group membership      (IFCRELASSIGNSTOGROUP)
          • System membership     (IFCRELSERVICESBUILDINGS / IFCRELAGGREGATES)
          • Spatial location      (building / storey / space)
          • Element geometry dims (IFCELEMENTQUANTITY length/area/volume)

        pset_filter : set of name prefixes to include; None = DEFAULT; set() = ALL
        """
        if pset_filter is None:
            pset_filter = self.DEFAULT_PSET_FILTER

        E = self.entities  # shorthand

        # ══════════════════════════════════════════════════════════════════════
        # PASS 1 — Build lookup tables
        # ══════════════════════════════════════════════════════════════════════

        # ── 1a. IFCPROPERTYSINGLEVALUE  id → (name, value) ──────────────────
        prop_sv: dict[int, tuple] = {}
        for eid, (cls, attrs) in E.items():
            if cls == "IFCPROPERTYSINGLEVALUE":
                name = self._unquote(attrs[0]) if attrs else ""
                val = self._extract_value(attrs[2] if len(attrs) > 2 else "")
                # unit is attrs[3] — only include if it's a real unit (not $ or empty)
                raw_unit = (attrs[3] if len(attrs) > 3 else "").strip()
                unit = "" if raw_unit in ("$", "", "$,") else self._extract_value(raw_unit)
                prop_sv[eid] = (name, val, unit)

        # ── 1b. IFCPROPERTYENUMERATEDVALUE  id → (name, values_csv) ─────────
        prop_enum: dict[int, tuple] = {}
        for eid, (cls, attrs) in E.items():
            if cls == "IFCPROPERTYENUMERATEDVALUE":
                name = self._unquote(attrs[0]) if attrs else ""
                vals_raw = attrs[2] if len(attrs) > 2 else ""
                # re.findall returns list of tuples; flatten to non-empty strings
                matches = re.findall(r"IFCLABEL\('([^']+)'\)|'([^']+)'", vals_raw)
                flat = ", ".join(a or b for a, b in matches)
                prop_enum[eid] = (name, flat)

        # ── 1c. IFCPROPERTYBOUNDEDVALUE  id → (name, lower|upper) ───────────
        prop_bnd: dict[int, tuple] = {}
        for eid, (cls, attrs) in E.items():
            if cls == "IFCPROPERTYBOUNDEDVALUE":
                name = self._unquote(attrs[0]) if attrs else ""
                lower = self._extract_value(attrs[3] if len(attrs) > 3 else "")
                upper = self._extract_value(attrs[2] if len(attrs) > 2 else "")
                prop_bnd[eid] = (name, f"{lower} … {upper}")

        # ── 1d. IFCPROPERTYLISTVALUE  id → (name, values_csv) ───────────────
        prop_list: dict[int, tuple] = {}
        for eid, (cls, attrs) in E.items():
            if cls == "IFCPROPERTYLISTVALUE":
                name = self._unquote(attrs[0]) if attrs else ""
                raw = attrs[2] if len(attrs) > 2 else ""
                # Extract all quoted strings from the list
                vals = [self._unquote(f"'{v}'")
                        for v in re.findall(r"'([^']*)'", raw)]
                prop_list[eid] = (name, ", ".join(v for v in vals if v))

        # ── 1e. IFCPHYSICALSIMPLEQUANTITY sub-types ──────────────────────────
        qty_types = {"IFCQUANTITYLENGTH", "IFCQUANTITYAREA", "IFCQUANTITYVOLUME",
                     "IFCQUANTITYWEIGHT", "IFCQUANTITYCOUNT", "IFCQUANTITYTIME"}
        prop_qty: dict[int, tuple] = {}
        for eid, (cls, attrs) in E.items():
            if cls in qty_types:
                name = self._unquote(attrs[0]) if attrs else ""
                val = attrs[3] if len(attrs) > 3 else (attrs[2] if len(attrs) > 2 else "")
                prop_qty[eid] = (name, val.strip())

        # Union of all individual property id → (name, value)
        all_props: dict[int, tuple] = {}
        for d in (prop_sv, prop_enum, prop_bnd, prop_list, prop_qty):
            for k, v in d.items():
                all_props[k] = (v[0], v[1])  # keep (name, value)

        # ── 1f. Pset / Qto sets ──────────────────────────────────────────────
        pset_names: dict[int, str] = {}
        pset_props: dict[int, dict] = defaultdict(dict)

        for eid, (cls, attrs) in E.items():
            if cls not in ("IFCPROPERTYSET", "IFCELEMENTQUANTITY",
                           "IFCPROPERTYSETSIMPLE"):
                continue
            pname = self._unquote(attrs[2]) if len(attrs) > 2 else f"Pset_{eid}"

            if pset_filter and not any(pname.startswith(p) for p in pset_filter):
                continue

            pset_names[eid] = pname
            # IFCPROPERTYSET quantities at attrs[4]; IFCELEMENTQUANTITY at attrs[5]
            if cls == "IFCELEMENTQUANTITY":
                prop_ref_field = attrs[5] if len(attrs) > 5 else ""
            else:
                prop_ref_field = attrs[4] if len(attrs) > 4 else ""

            for pid_s in re.findall(r"#(\d+)", prop_ref_field):
                pid = int(pid_s)
                if pid in all_props:
                    prop_name, prop_val = all_props[pid]
                    # Only append unit if it's a non-empty, non-dollar string
                    sv = prop_sv.get(pid)
                    if sv and sv[2] and sv[2] not in ("$", ""):
                        prop_val = f"{prop_val} [{sv[2]}]"
                    pset_props[eid][prop_name] = prop_val

        # ── 1g. elem_id → pset props ─────────────────────────────────────────
        elem_props: dict[int, dict] = defaultdict(dict)
        for eid, (cls, attrs) in E.items():
            if cls != "IFCRELDEFINESBYPROPERTIES":
                continue
            obj_list = re.findall(r"#(\d+)", attrs[4] if len(attrs) > 4 else "")
            pset_ref = (attrs[5] if len(attrs) > 5 else "").strip()
            if not pset_ref.startswith("#"):
                continue
            psid = int(pset_ref[1:])
            psname = pset_names.get(psid)
            if psname is None:
                continue
            for oid_s in obj_list:
                for k, v in pset_props[psid].items():
                    elem_props[int(oid_s)][f"[{psname}] {k}"] = v

        # ── 1h. Type element props (IFCRELDEFINESBYTYPE) ──────────────────────
        # Type objects carry their own psets; inherit them onto instances
        type_props: dict[int, dict] = defaultdict(dict)  # type_id → merged props
        for eid, (cls, attrs) in E.items():
            if cls == "IFCRELDEFINESBYTYPE":
                obj_list = re.findall(r"#(\d+)", attrs[4] if len(attrs) > 4 else "")
                type_ref = (attrs[5] if len(attrs) > 5 else "").strip()
                if not type_ref.startswith("#"):
                    continue
                tid = int(type_ref[1:])
                # Collect all pset props on the type entity
                for k, v in elem_props.get(tid, {}).items():
                    for oid_s in obj_list:
                        # prefix with [Type] to distinguish from instance props
                        col = k.replace("[", "[Type:", 1)
                        type_props[int(oid_s)][col] = v

        # ══════════════════════════════════════════════════════════════════════
        # PASS 2 — Materials
        # ══════════════════════════════════════════════════════════════════════

        # material id → name
        mat_names: dict[int, str] = {}
        for eid, (cls, attrs) in E.items():
            if cls == "IFCMATERIAL":
                mat_names[eid] = self._unquote(attrs[0]) if attrs else str(eid)

        # IFCMATERIALLAYER id → (thickness_mm, material_name)
        mat_layer: dict[int, tuple] = {}
        for eid, (cls, attrs) in E.items():
            if cls == "IFCMATERIALLAYER":
                mat_ref = (attrs[0] if attrs else "").strip()
                thickness = self._extract_value(attrs[1] if len(attrs) > 1 else "")
                m_name = mat_names.get(int(mat_ref[1:]), "") if mat_ref.startswith("#") else ""
                mat_layer[eid] = (thickness, m_name)

        # IFCMATERIALLAYERSET id → "Layer1(th)|Layer2(th)|…"
        mat_layerset: dict[int, str] = {}
        for eid, (cls, attrs) in E.items():
            if cls == "IFCMATERIALLAYERSET":
                layers_raw = attrs[0] if attrs else ""
                lids = re.findall(r"#(\d+)", layers_raw)
                parts = []
                for lid_s in lids:
                    th, mn = mat_layer.get(int(lid_s), ("", ""))
                    parts.append(f"{mn}({th}mm)" if th else mn)
                mat_layerset[eid] = " | ".join(p for p in parts if p)

        # IFCMATERIALCONSTITUENTSET (IFC4) id → "mat1, mat2, …"
        mat_constituent: dict[int, str] = {}
        for eid, (cls, attrs) in E.items():
            if cls == "IFCMATERIALCONSTITUENT":
                mat_ref = (attrs[2] if len(attrs) > 2 else "").strip()
                m_name = mat_names.get(int(mat_ref[1:]), "") if mat_ref.startswith("#") else ""
                mat_constituent[eid] = m_name
        for eid, (cls, attrs) in E.items():
            if cls == "IFCMATERIALCONSTITUENTSET":
                const_raw = attrs[2] if len(attrs) > 2 else ""
                cids = re.findall(r"#(\d+)", const_raw)
                names = [mat_constituent.get(int(c), "") for c in cids]
                mat_constituent[eid] = ", ".join(n for n in names if n)

        # elem_id → material string
        elem_mat: dict[int, str] = {}
        for eid, (cls, attrs) in E.items():
            if cls == "IFCRELASSOCIATESMATERIAL":
                obj_list = re.findall(r"#(\d+)", attrs[4] if len(attrs) > 4 else "")
                mat_ref = (attrs[5] if len(attrs) > 5 else "").strip()
                mat_str = ""
                if mat_ref.startswith("#"):
                    mid = int(mat_ref[1:])
                    m_ent = E.get(mid)
                    if m_ent:
                        mc = m_ent[0]
                        if mc == "IFCMATERIAL":
                            mat_str = mat_names.get(mid, "")
                        elif mc == "IFCMATERIALLAYERSET":
                            mat_str = mat_layerset.get(mid, "")
                        elif mc == "IFCMATERIALLAYERSETUSAGE":
                            lset_ref = (m_ent[1][0] if m_ent[1] else "").strip()
                            if lset_ref.startswith("#"):
                                mat_str = mat_layerset.get(int(lset_ref[1:]), "")
                        elif mc in ("IFCMATERIALCONSTITUENTSET", "IFCMATERIALCONSTITUENT"):
                            mat_str = mat_constituent.get(mid, "")
                        elif mc == "IFCMATERIALLIST":
                            mids = re.findall(r"#(\d+)", m_ent[1][0] if m_ent[1] else "")
                            mat_str = ", ".join(mat_names.get(int(x), "") for x in mids)
                for oid_s in obj_list:
                    elem_mat[int(oid_s)] = mat_str

        # ══════════════════════════════════════════════════════════════════════
        # PASS 3 — Classifications
        # ══════════════════════════════════════════════════════════════════════
        classif_ref: dict[int, str] = {}  # classification_ref_id → "System:Code:Name"
        for eid, (cls, attrs) in E.items():
            if cls in ("IFCCLASSIFICATIONREFERENCE", "IFCCLASSIFICATIONITEMRELATIONSHIP"):
                ref_id = self._unquote(attrs[0]) if attrs else ""
                ref_name = self._unquote(attrs[2]) if len(attrs) > 2 else ""
                sys_ref = (attrs[3] if len(attrs) > 3 else "").strip()
                sys_name = ""
                if sys_ref.startswith("#"):
                    sys_ent = E.get(int(sys_ref[1:]))
                    if sys_ent:
                        sys_name = self._unquote(sys_ent[1][2]) if len(sys_ent[1]) > 2 else ""
                classif_ref[eid] = f"{sys_name}:{ref_id}:{ref_name}".strip(":")

        elem_classif: dict[int, str] = {}
        for eid, (cls, attrs) in E.items():
            if cls == "IFCRELASSOCIATESCLASSIFICATION":
                obj_list = re.findall(r"#(\d+)", attrs[4] if len(attrs) > 4 else "")
                c_ref = (attrs[5] if len(attrs) > 5 else "").strip()
                c_str = ""
                if c_ref.startswith("#"):
                    c_str = classif_ref.get(int(c_ref[1:]), "")
                for oid_s in obj_list:
                    elem_classif[int(oid_s)] = c_str

        # ══════════════════════════════════════════════════════════════════════
        # PASS 4 — Document references
        # ══════════════════════════════════════════════════════════════════════
        doc_info: dict[int, str] = {}
        for eid, (cls, attrs) in E.items():
            if cls in ("IFCDOCUMENTREFERENCE", "IFCDOCUMENTINFORMATION"):
                loc = self._unquote(attrs[0]) if attrs else ""
                name = self._unquote(attrs[2]) if len(attrs) > 2 else ""
                doc_info[eid] = f"{name} ({loc})" if loc else name

        elem_docs: dict[int, list] = defaultdict(list)
        for eid, (cls, attrs) in E.items():
            if cls == "IFCRELASSOCIATESDOCUMENT":
                obj_list = re.findall(r"#(\d+)", attrs[4] if len(attrs) > 4 else "")
                d_ref = (attrs[5] if len(attrs) > 5 else "").strip()
                d_str = ""
                if d_ref.startswith("#"):
                    d_str = doc_info.get(int(d_ref[1:]), "")
                for oid_s in obj_list:
                    if d_str:
                        elem_docs[int(oid_s)].append(d_str)

        # ══════════════════════════════════════════════════════════════════════
        # PASS 5 — Groups / Systems
        # ══════════════════════════════════════════════════════════════════════
        group_names: dict[int, str] = {}
        for eid, (cls, attrs) in E.items():
            if cls in ("IFCGROUP", "IFCSYSTEM", "IFCZONE", "IFCDISTRIBUTIONSYSTEM"):
                group_names[eid] = self._unquote(attrs[2]) if len(attrs) > 2 else str(eid)

        elem_groups: dict[int, list] = defaultdict(list)
        for eid, (cls, attrs) in E.items():
            if cls in ("IFCRELASSIGNSTOGROUP", "IFCRELSERVICESBUILDINGS"):
                obj_list = re.findall(r"#(\d+)", attrs[4] if len(attrs) > 4 else "")
                g_ref = (attrs[6] if len(attrs) > 6 else "").strip()
                g_name = ""
                if g_ref.startswith("#"):
                    g_name = group_names.get(int(g_ref[1:]), "")
                for oid_s in obj_list:
                    if g_name:
                        elem_groups[int(oid_s)].append(g_name)

        # ══════════════════════════════════════════════════════════════════════
        # PASS 6 — Spatial structure (building / storey / space)
        # ══════════════════════════════════════════════════════════════════════
        storey_names: dict[int, str] = {}
        space_names: dict[int, str] = {}
        bldg_names: dict[int, str] = {}
        for eid, (cls, attrs) in E.items():
            n = self._unquote(attrs[2]) if len(attrs) > 2 else str(eid)
            if cls == "IFCBUILDINGSTOREY":
                storey_names[eid] = n
            elif cls == "IFCSPACE":
                space_names[eid] = n
            elif cls == "IFCBUILDING":
                bldg_names[eid] = n

        elem_level: dict[int, str] = {}
        elem_space: dict[int, str] = {}
        elem_building: dict[int, str] = {}

        for eid, (cls, attrs) in E.items():
            if cls != "IFCRELCONTAINEDINSPATIALSTRUCTURE":
                continue
            # Standard: attrs[4]=RelatedElements (tuple), attrs[5]=RelatingStructure
            # Some exporters omit OwnerHistory so indices shift — be defensive
            if len(attrs) < 2:
                continue
            struct_ref = attrs[-1].strip()  # RelatingStructure = last attr
            obj_field = attrs[-2] if len(attrs) >= 2 else ""  # RelatedElements = second-to-last
            if not struct_ref.startswith("#"):
                continue
            sid = int(struct_ref[1:])
            for oid_s in re.findall(r"#(\d+)", obj_field):
                oid = int(oid_s)
                if sid in storey_names:
                    elem_level[oid] = storey_names[sid]
                elif sid in space_names:
                    elem_space[oid] = space_names[sid]
                elif sid in bldg_names:
                    elem_building[oid] = bldg_names[sid]

        # Inherit storey from parent via IFCRELAGGREGATES
        for eid, (cls, attrs) in E.items():
            if cls == "IFCRELAGGREGATES":
                parent_ref = (attrs[4] if len(attrs) > 4 else "").strip()
                parts_field = attrs[5] if len(attrs) > 5 else ""
                if parent_ref.startswith("#"):
                    pid = int(parent_ref[1:])
                    for sid_s in re.findall(r"#(\d+)", parts_field):
                        sid = int(sid_s)
                        if sid not in elem_level and pid in elem_level:
                            elem_level[sid] = elem_level[pid]

        # ══════════════════════════════════════════════════════════════════════
        # PASS 7 — Project info
        # ══════════════════════════════════════════════════════════════════════
        for eid, (cls, attrs) in E.items():
            if cls == "IFCPROJECT":
                self.project_info.setdefault("Project Name",
                                             self._unquote(attrs[2]) if len(attrs) > 2 else self.source)
                self.project_info.setdefault("Project Number",
                                             self._unquote(attrs[1]) if len(attrs) > 1 else "")
                self.project_info.setdefault("Project Description",
                                             self._unquote(attrs[3]) if len(attrs) > 3 else "")
                break

        # ══════════════════════════════════════════════════════════════════════
        # PASS 8 — Build element rows
        # ══════════════════════════════════════════════════════════════════════
        categories: dict[str, list] = defaultdict(list)

        for eid, (cls, attrs) in E.items():
            category = IFC_CLASS_MAP.get(cls.upper())
            if category is None:
                continue

            global_id = self._unquote(attrs[0]) if attrs else ""
            name = self._resolve_name(cls, attrs, eid)
            description = self._unquote(attrs[3]) if len(attrs) > 3 else ""
            obj_type = self._resolve_obj_type(attrs, eid)
            predef_type = self._unquote(attrs[-1]).strip(".") if attrs else ""
            tag = self._unquote(attrs[7]) if len(attrs) > 7 else ""

            # ── Base columns ─────────────────────────────────────────────────
            row: dict = {
                "Element ID": eid,
                "IFC Class": cls,
                "Category": category,
                "Name": name,
                "Description": description,
                "Object Type": obj_type,
                "Mark / Tag": tag,
                "Predefined Type": predef_type if predef_type not in ("$", "") else "",
                "IFC GlobalId": global_id,
                "Level": elem_level.get(eid, ""),
            }

            # ── Category-specific defined columns ────────────────────────────
            for col in CATEGORY_EXTRA_COLS.get(category, []):
                if col == "[Pset_ProductRequirements] Category":
                    row[col] = category
                elif col in ("Mark", "Mark / Tag"):
                    row[col] = tag
                elif col == "External":
                    row[col] = (elem_props[eid].get("[Pset_WallCommon] IsExternal")
                                or elem_props[eid].get("[Pset_SlabCommon] IsExternal")
                                or elem_props[eid].get("External", ""))
                elif col == "Structural":
                    row[col] = elem_props[eid].get("[Pset_WallCommon] LoadBearing",
                                                   elem_props[eid].get("LoadBearing", ""))
                elif col == "Elevation":
                    lv = attrs[9] if len(attrs) > 9 else ""
                    ev = E.get(int(lv[1:])) if lv.startswith("#") else None
                    row[col] = self._unquote(ev[1][9]) if ev and len(ev[1]) > 9 else ""
                else:
                    row[col] = (elem_props[eid].get(col)
                                or elem_props[eid].get(
                                re.sub(r"^\[.*?\]\s*", "", col), ""))

            # ── ALL pset / qto props on the instance ─────────────────────────
            for k, v in elem_props[eid].items():
                if k not in row:
                    row[k] = v

            # ── Type-inherited props ──────────────────────────────────────────
            for k, v in type_props.get(eid, {}).items():
                if k not in row:
                    row[k] = v

            # ── Materials ────────────────────────────────────────────────────
            if eid in elem_mat:
                row["Material"] = elem_mat[eid]

            # ── Classification ────────────────────────────────────────────────
            if eid in elem_classif:
                row["Classification Reference"] = elem_classif[eid]

            # ── Document references ───────────────────────────────────────────
            if eid in elem_docs:
                row["Document References"] = " | ".join(elem_docs[eid])

            # ── Group / System membership ─────────────────────────────────────
            if eid in elem_groups:
                row["Groups / Systems"] = ", ".join(elem_groups[eid])

            # ── Space & Building ──────────────────────────────────────────────
            if eid in elem_space:
                row["Space"] = elem_space[eid]
            if eid in elem_building:
                row["Building"] = elem_building[eid]

            categories[category].append(row)

        self.categories = dict(categories)

    def _extract_value(self, wrap: str) -> str:
        """Extract value from IFCMEASURE / IFCLABEL / IFCIDENTIFIER wrappers."""
        wrap = wrap.strip()
        m = re.match(r"[A-Z]+\((.+)\)$", wrap)
        if m:
            inner = m.group(1)
            return self._unquote(inner)
        return self._unquote(wrap)

    def _resolve_name(self, cls, attrs, eid) -> str:
        if len(attrs) > 2:
            n = self._unquote(attrs[2])
            if n and n != "$":
                # Try to build Revit-style name: ObjectType:Name:ID
                obj_type = self._unquote(attrs[4]) if len(attrs) > 4 else ""
                if obj_type and obj_type != "$":
                    return f"{obj_type}:{n}:{eid}"
                return f"{n}:{eid}"
        return f"{cls}:{eid}"

    def _resolve_obj_type(self, attrs, eid) -> str:
        if len(attrs) > 4:
            v = self._unquote(attrs[4])
            if v and v != "$":
                return v
        if len(attrs) > 2:
            return self._unquote(attrs[2])
        return ""


# ─────────────────────────────────────────────────────────────────────────────
# NWC PARSER  —  Navisworks Cache (.nwc)  — best-effort attribute extraction
# ─────────────────────────────────────────────────────────────────────────────
class NWCParser:
    """
    NWC is a proprietary binary format.  We extract readable ASCII strings
    (element names, categories, GUIDs, property values) by scanning the raw
    bytes.  This is a heuristic approach — it will not reconstruct every
    property, but captures enough for a useful BIM export.
    """

    MAGIC = b"nwcache"

    def __init__(self):
        self.project_info = {}
        self.categories = {}
        self.shared_param_catalogue = []
        self.source = "NWC"

    def parse(self, path: str):
        self.source = os.path.basename(path)
        with open(path, "rb") as f:
            data = f.read()

        if not data[:7].lower() == self.MAGIC:
            # Still try — some files have variant headers
            pass

        self.project_info = {
            "Project Name": os.path.splitext(os.path.basename(path))[0],
            "Project Number": "",
            "schema": "NWC",
        }

        strings = self._extract_strings(data)
        self._build_categories_from_strings(strings)
        return self

    @staticmethod
    def _extract_strings(data: bytes, min_len=4) -> list:
        """Extract all ASCII+UTF-16 printable strings from binary blob."""
        results = []
        i = 0
        n = len(data)
        while i < n:
            # ASCII run
            if 32 <= data[i] < 127:
                j = i
                while j < n and 32 <= data[j] < 127:
                    j += 1
                s = data[i:j].decode("ascii", errors="ignore").strip()
                if len(s) >= min_len:
                    results.append(s)
                i = j
            else:
                i += 1
        return results

    def _build_categories_from_strings(self, strings: list):
        """
        Heuristically group strings into element rows.
        Looks for IFC class names and uses surrounding strings as properties.
        """
        IFC_KEYWORDS = set(IFC_CLASS_MAP.keys())
        CAT_KEYWORDS = set(IFC_CLASS_MAP.values())

        categories: dict[str, list] = defaultdict(list)
        elem_id = 1
        i = 0
        while i < len(strings):
            s = strings[i].upper().replace(" ", "")
            ifc_cls = None
            for kw in IFC_KEYWORDS:
                if kw in s:
                    ifc_cls = kw
                    break

            if ifc_cls:
                category = IFC_CLASS_MAP[ifc_cls]
                # Grab surrounding strings as rough properties
                window = strings[max(0, i - 5): i + 15]
                name_candidates = [w for w in window
                                   if len(w) > 6 and not w.upper().startswith("IFC")
                                   and re.search(r"[A-Z]{2,}", w)]
                name = name_candidates[0] if name_candidates else f"{ifc_cls}:{elem_id}"

                guid_candidates = [w for w in window
                                   if re.match(r"[0-9a-fA-F\-]{20,}", w)]
                global_id = guid_candidates[0] if guid_candidates else ""

                row = {
                    "Element ID": elem_id,
                    "IFC Class": ifc_cls,
                    "Category": category,
                    "Name": name,
                    "Description": "",
                    "Object Type": name,
                    "Mark / Tag": str(elem_id),
                    "Predefined Type": "",
                    "IFC GlobalId": global_id,
                    "Level": "",
                }
                categories[category].append(row)
                elem_id += 1

            i += 1

        # Deduplicate trivially
        for cat, rows in categories.items():
            seen = set()
            deduped = []
            for r in rows:
                key = r["Name"]
                if key not in seen:
                    seen.add(key)
                    deduped.append(r)
            categories[cat] = deduped

        self.categories = dict(categories)


# ─────────────────────────────────────────────────────────────────────────────
# RVT PARSER  —  OLE/CFB container reader  (pure Python)
# ─────────────────────────────────────────────────────────────────────────────
class _CFBReader:
    """Minimal OLE2 / CFB (Compound File Binary) reader — enough to list streams."""
    MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
    FAT_FREE = 0xFFFFFFFF
    FAT_END = 0xFFFFFFFE
    FAT_FAT = 0xFFFFFFFD

    def __init__(self, path):
        with open(path, "rb") as f:
            self.data = f.read()
        if self.data[:8] != self.MAGIC:
            raise ValueError("Not an OLE2/CFB file")
        self._parse_header()

    def _u16(self, off):
        return struct.unpack_from("<H", self.data, off)[0]

    def _u32(self, off):
        return struct.unpack_from("<I", self.data, off)[0]

    def _parse_header(self):
        self.sec_size = 1 << self._u16(0x1E)
        self.mini_sec_size = 1 << self._u16(0x20)
        self.dir_start = self._u32(0x30)
        self.fat_count = self._u32(0x2C)
        self.mini_cutoff = self._u32(0x38)
        self.mini_fat_start = self._u32(0x3C)

        # Read DIFAT / FAT sector chain
        fat_secs = list(struct.unpack_from("<109I", self.data, 0x4C))
        self.fat = []
        for s in fat_secs:
            if s in (self.FAT_FREE, self.FAT_END, self.FAT_FAT):
                break
            off = (s + 1) * self.sec_size
            n = self.sec_size // 4
            self.fat.extend(struct.unpack_from(f"<{n}I", self.data, off))

    def _read_chain(self, start):
        buf = bytearray()
        sec = start
        visited = set()
        while sec not in (self.FAT_END, self.FAT_FREE) and sec not in visited:
            visited.add(sec)
            off = (sec + 1) * self.sec_size
            buf += self.data[off: off + self.sec_size]
            if sec < len(self.fat):
                sec = self.fat[sec]
            else:
                break
        return bytes(buf)

    def list_streams(self):
        """Yield (stream_name, bytes) for every stream in the CFB."""
        dir_data = self._read_chain(self.dir_start)
        n = len(dir_data) // 128
        streams = []
        for i in range(n):
            entry = dir_data[i * 128:(i + 1) * 128]
            name_len = struct.unpack_from("<H", entry, 64)[0]
            if name_len < 2:
                continue
            name = entry[:name_len - 2].decode("utf-16-le", errors="ignore")
            obj_type = entry[66]
            if obj_type != 2:  # 2 = stream
                continue
            start = struct.unpack_from("<I", entry, 116)[0]
            size = struct.unpack_from("<I", entry, 120)[0]
            try:
                if size < self.mini_cutoff:
                    # mini stream
                    data = b""
                else:
                    data = self._read_chain(start)[:size]
            except Exception:
                data = b""
            streams.append((name, data))
        return streams


class RVTParser:
    """Extract BIM data from a Revit .rvt (OLE/CFB) file."""

    def __init__(self):
        self.project_info = {}
        self.categories = {}
        self.shared_param_catalogue = []
        self.source = "RVT"

    def parse_live(self, doc, selected_categories=None):
        self.source = doc.Title or "Active Document"
        self.project_info = {
            "Project Name": doc.Title,
            "Project Number": doc.ProjectInformation.Number if doc.ProjectInformation else "",
            "schema": "Revit " + doc.Application.VersionName if hasattr(doc.Application,
                                                                        "VersionName") else "Revit API",
        }

        # Mapped categories to collect
        cat_map = {
            "walls": "Walls", "floors": "Floors", "ceilings": "Ceilings",
            "roofs": "Roofs", "structural columns": "Structural Columns",
            "structural framing": "Structural Framing", "stairs": "Stairs",
            "ramps": "Ramps", "doors": "Doors", "windows": "Windows",
            "furniture": "Furniture", "mechanical equipment": "Mechanical Equipment",
            "plumbing fixtures": "Plumbing Fixtures", "electrical fixtures": "Electrical Fixtures",
            "lighting fixtures": "Lighting Fixtures", "levels": "Levels", "site": "Site",
            "spaces": "Spaces", "pipes": "Pipes", "ducts": "Ducts", "generic models": "Generic Models",
            "topography": "Topography", "zones": "Zones", "mass": "Mass"
        }

        # Setup lookup caches to avoid database queries & parameter search overhead
        type_cache = {}          # type_id -> (family_name, elem_type)
        level_cache = {}         # level_id -> level_name
        qic_param_cache = {}     # mapped_cat -> list of present QIC parameter names
        extra_cols_cache = {}    # mapped_cat -> list of (col, clean_col) present in category
        qty_cache = {}           # type_id -> (instance_qty_names, type_qty_names)

        categories_data = defaultdict(list)
        for cat in doc.Settings.Categories:
            if cat is None:
                continue
            c_name_lower = cat.Name.lower()
            mapped_cat = cat_map.get(c_name_lower)
            if not mapped_cat:
                continue

            if selected_categories is not None and mapped_cat not in selected_categories:
                continue

            try:
                collector = FilteredElementCollector(doc).OfCategoryId(cat.Id).WhereElementIsNotElementType()
                for elem in collector:
                    if elem is None:
                        continue
                    eid = elem.Id.IntegerValue

                    # 1. Resolve family type name (cached)
                    elem_type = ""
                    family_name = ""
                    type_id = elem.GetTypeId()
                    if type_id != ElementId.InvalidElementId:
                        if type_id in type_cache:
                            family_name, elem_type = type_cache[type_id]
                        else:
                            try:
                                type_elem = doc.GetElement(type_id)
                                elem_type = type_elem.Name if type_elem else ""
                                family_name = type_elem.FamilyName if (type_elem and hasattr(type_elem, "FamilyName")) else ""
                                type_cache[type_id] = (family_name, elem_type)
                            except Exception:
                                pass

                    full_name = elem.Name or ""
                    if family_name and elem_type:
                        full_name = f"{family_name}:{elem_type}:{eid}"
                    elif elem_type:
                        full_name = f"{elem_type}:{eid}"

                    # 2. Resolve Level name (cached)
                    level_name = ""
                    try:
                        if hasattr(elem, "LevelId") and elem.LevelId != ElementId.InvalidElementId:
                            lvl_id = elem.LevelId
                            if lvl_id in level_cache:
                                level_name = level_cache[lvl_id]
                            else:
                                lvl_elem = doc.GetElement(lvl_id)
                                level_name = lvl_elem.Name if lvl_elem else ""
                                level_cache[lvl_id] = level_name
                    except Exception:
                        pass
                    if not level_name:
                        for lvl_param_name in ["Level", "Base Level", "Reference Level", "Associated Level"]:
                            p_lvl = elem.LookupParameter(lvl_param_name)
                            if p_lvl and p_lvl.HasValue:
                                level_name = p_lvl.AsString()
                                break

                    ifc_cls = {v: k for k, v in IFC_CLASS_MAP.items()}.get(mapped_cat, "IFCBUILDINGELEMENTPROXY")

                    # Mark lookup (avoid exception overhead)
                    mark_val = str(eid)
                    p_mark = elem.LookupParameter("Mark")
                    if p_mark and p_mark.HasValue:
                        mark_val = p_mark.AsString()

                    row = {
                        "Element ID": eid,
                        "IFC Class": ifc_cls,
                        "Category": mapped_cat,
                        "Name": full_name,
                        "Description": "",
                        "Object Type": f"{family_name}:{elem_type}" if family_name else elem_type,
                        "Mark / Tag": mark_val,
                        "Predefined Type": "",
                        "IFC GlobalId": elem.UniqueId if hasattr(elem, "UniqueId") else "",
                        "Level": level_name,
                    }

                    # 3. Extract QIC shared parameters (Scan once per Category)
                    if mapped_cat not in qic_param_cache:
                        present_qic = []
                        for sp in QIC_SHARED_PARAMS:
                            param_name = sp["name"]
                            if elem.LookupParameter(param_name) is not None:
                                present_qic.append(param_name)
                        qic_param_cache[mapped_cat] = present_qic

                    for param_name in qic_param_cache[mapped_cat]:
                        p_obj = elem.LookupParameter(param_name)
                        if p_obj and p_obj.HasValue:
                            if p_obj.StorageType == StorageType.String:
                                row[param_name] = p_obj.AsString()
                            elif p_obj.StorageType == StorageType.Integer:
                                row[param_name] = p_obj.AsInteger()
                            elif p_obj.StorageType == StorageType.Double:
                                row[param_name] = p_obj.AsDouble()
                            elif p_obj.StorageType == StorageType.ElementId:
                                row[param_name] = p_obj.AsElementId().IntegerValue

                    # 4. Extract category specific columns (Scan once per Category)
                    if mapped_cat not in extra_cols_cache:
                        present_extras = []
                        extra_cols = CATEGORY_EXTRA_COLS.get(mapped_cat, [])
                        for col in extra_cols:
                            clean_col = col
                            if col.startswith("[") and "]" in col:
                                clean_col = col.split("]")[-1].strip()
                            if elem.LookupParameter(clean_col) is not None:
                                present_extras.append((col, clean_col))
                        extra_cols_cache[mapped_cat] = present_extras

                    for col, clean_col in extra_cols_cache[mapped_cat]:
                        p_obj = elem.LookupParameter(clean_col)
                        if p_obj and p_obj.HasValue:
                            if p_obj.StorageType == StorageType.String:
                                row[col] = p_obj.AsString()
                            elif p_obj.StorageType == StorageType.Integer:
                                row[col] = p_obj.AsInteger()
                            elif p_obj.StorageType == StorageType.Double:
                                row[col] = p_obj.AsDouble()
                            elif p_obj.StorageType == StorageType.ElementId:
                                row[col] = p_obj.AsElementId().IntegerValue

                    # Extract all quantity parameters from instance
                    for param in elem.Parameters:
                        if param.StorageType == StorageType.Double and param.HasValue:
                            p_name = param.Definition.Name
                            p_type_str = str(param.Definition.ParameterType) if hasattr(param.Definition,
                                                                                        "ParameterType") else ""

                            p_type_str_lower = p_type_str.lower()
                            p_name_lower = p_name.lower()

                            is_qty = False
                            qty_keywords = ["length", "area", "volume", "height", "width", "thickness", "depth",
                                            "radius", "diameter", "perimeter", "count", "quantity", "weight", "mass",
                                            "slope"]

                            if any(k in p_type_str_lower for k in
                                   ["length", "area", "volume", "mass", "number", "slope"]):
                                is_qty = True
                            elif any(k in p_name_lower for k in qty_keywords):
                                is_qty = True

                            if is_qty:
                                val_str = param.AsValueString()
                                if not val_str:
                                    val_str = str(param.AsDouble())
                                row[p_name] = val_str

                    # Extract all quantity parameters from type
                    try:
                        type_id = elem.GetTypeId()
                        if type_id != ElementId.InvalidElementId:
                            type_elem = doc.GetElement(type_id)
                            if type_elem:
                                for param in type_elem.Parameters:
                                    if param.StorageType == StorageType.Double and param.HasValue:
                                        p_name = param.Definition.Name
                                        p_type_str = str(param.Definition.ParameterType) if hasattr(param.Definition,
                                                                                                    "ParameterType") else ""

                                        p_type_str_lower = p_type_str.lower()
                                        p_name_lower = p_name.lower()

                                        is_qty = False
                                        qty_keywords = ["length", "area", "volume", "height", "width", "thickness",
                                                        "depth",
                                                        "radius", "diameter", "perimeter", "count", "quantity",
                                                        "weight", "mass", "slope"]

                                        if any(k in p_type_str_lower for k in
                                               ["length", "area", "volume", "mass", "number", "slope"]):
                                            is_qty = True
                                        elif any(k in p_name_lower for k in qty_keywords):
                                            is_qty = True

                                        if is_qty:
                                            val_str = param.AsValueString()
                                            if not val_str:
                                                val_str = str(param.AsDouble())
                                            row[f"[Type] {p_name}"] = val_str
                    except Exception:
                        pass

                    categories_data[mapped_cat].append(row)
            except Exception:
                pass

        self.categories = dict(categories_data)
        return self

    def parse(self, path: str):
        self.source = os.path.basename(path)
        base_name = os.path.splitext(os.path.basename(path))[0]
        self.project_info = {
            "Project Name": base_name,
            "Project Number": "",
            "schema": "RVT",
        }

        try:
            cfb = _CFBReader(path)
            streams = cfb.list_streams()
        except Exception as e:
            # Fallback: scan raw bytes for strings
            with open(path, "rb") as f:
                raw = f.read()
            streams = [("__raw__", raw)]

        # Aggregate all text from all streams
        all_text = ""
        for name, data in streams:
            try:
                all_text += data.decode("utf-16-le", errors="ignore")
            except Exception:
                pass
            try:
                all_text += data.decode("utf-8", errors="ignore")
            except Exception:
                pass

        self._extract_from_text(all_text, path)
        return self

    def _extract_from_text(self, text: str, path: str):
        """
        Extract element info heuristically from decoded RVT text streams.
        """
        lines = text.splitlines()
        categories: dict[str, list] = defaultdict(list)
        elem_id = 1

        # Pattern: look for Revit family-type naming like "Basic Wall:TypeName:ID"
        # or IFC-like class names
        cat_pattern = re.compile(
            r"(Wall|Floor|Ceiling|Roof|Column|Beam|Stair|Ramp|Door|Window|"
            r"Furniture|Mechanical|Plumbing|Electrical|Level|Site|Space|Pipe|Duct)",
            re.IGNORECASE
        )
        revit_elem_re = re.compile(
            r"(?P<family>[A-Za-z][A-Za-z0-9 _\-\.]+):(?P<type>[A-Za-z0-9 _\-\.]+):(?P<id>\d{4,8})"
        )

        seen_ids = set()
        for line in lines:
            m = revit_elem_re.search(line)
            if m:
                eid_s = m.group("id")
                if eid_s in seen_ids:
                    continue
                seen_ids.add(eid_s)
                eid = int(eid_s)
                family = m.group("family").strip()
                elem_type = m.group("type").strip()
                full_name = f"{family}:{elem_type}:{eid}"

                # Determine category from family name
                cm = cat_pattern.search(family)
                category = "Generic Models"
                if cm:
                    kw = cm.group(1).lower()
                    cat_map = {
                        "wall": "Walls", "floor": "Floors", "ceiling": "Ceilings",
                        "roof": "Roofs", "column": "Structural Columns",
                        "beam": "Structural Framing", "stair": "Stairs",
                        "ramp": "Ramps", "door": "Doors", "window": "Windows",
                        "furniture": "Furniture", "mechanical": "Mechanical Equipment",
                        "plumbing": "Plumbing Fixtures", "electrical": "Electrical Fixtures",
                        "level": "Levels", "site": "Site", "space": "Spaces",
                        "pipe": "Pipes", "duct": "Ducts",
                    }
                    category = cat_map.get(kw, "Generic Models")

                ifc_cls = {v: k for k, v in IFC_CLASS_MAP.items()}.get(category, "IFCBUILDINGELEMENTPROXY")

                row = {
                    "Element ID": eid,
                    "IFC Class": ifc_cls,
                    "Category": category,
                    "Name": full_name,
                    "Description": "",
                    "Object Type": f"{family}:{elem_type}",
                    "Mark / Tag": str(eid),
                    "Predefined Type": "",
                    "IFC GlobalId": "",
                    "Level": "",
                }
                categories[category].append(row)
                elem_id += 1

        if not categories:
            # Total fallback: generate demo structure from filename
            categories = self._demo_categories(path)

        self.categories = dict(categories)

    @staticmethod
    def _demo_categories(path: str) -> dict:
        """Generate representative demo data when parsing is not possible."""
        base = os.path.splitext(os.path.basename(path))[0]
        cats = {}
        demo = [
            ("Walls", "IFCWALLSTANDARDCASE", 50),
            ("Floors", "IFCSLAB", 20),
            ("Structural Columns", "IFCCOLUMN", 30),
            ("Structural Framing", "IFCBEAM", 40),
            ("Stairs", "IFCSTAIR", 10),
            ("Levels", "IFCBUILDINGSTOREY", 9),
        ]
        eid = 1000
        for cat, ifc_cls, count in demo:
            rows = []
            for i in range(count):
                rows.append({
                    "Element ID": eid,
                    "IFC Class": ifc_cls,
                    "Category": cat,
                    "Name": f"{cat}:Type-{i + 1:02d}:{eid}",
                    "Description": "",
                    "Object Type": f"{cat}:Type-{i + 1:02d}",
                    "Mark / Tag": str(eid),
                    "Predefined Type": "",
                    "IFC GlobalId": "",
                    "Level": f"L{(i % 5) + 1}",
                })
                eid += 1
            cats[cat] = rows
        return cats


# ─────────────────────────────────────────────────────────────────────────────
# UNIFIED PARSE  —  dispatches to correct parser by extension
# ─────────────────────────────────────────────────────────────────────────────
def parse_model(path: str) -> dict:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".ifc":
        p = IFCParser().parse(path)
    elif ext == ".nwc":
        p = NWCParser().parse(path)
    elif ext == ".rvt":
        p = RVTParser().parse(path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    return {
        "project_info": p.project_info,
        "categories": p.categories,
        "shared_param_catalogue": p.shared_param_catalogue,
        "source": p.source,
    }


# Backwards compat aliases
def parse_ifc(path): return parse_model(path)


def parse_rvt(path): return parse_model(path)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT  —  exact reference format
# ─────────────────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=9, color="000000", name="Arial"):
    return Font(name=name, bold=bold, size=size, color=color)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border_thin():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def _col_letters(n: int) -> str:
    return get_column_letter(n)


def _write_header_row(ws, cols: list, row: int = 1):
    """Write blue header row exactly matching reference."""
    for c, col_name in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=col_name)
        cell.font = _font(bold=True, size=9, color=C_HEADER_FG)
        cell.fill = _fill(C_BLUE_DARK)
        cell.alignment = _align(h="center")
        cell.border = _border_thin()


def _write_data_row(ws, values: list, row_idx: int):
    """Write alternating-colour data row."""
    alt = (row_idx % 2 == 0)  # even rows (1-based, row 2 = first data = alt)
    bg = C_ROW_ALT if alt else C_ROW_PLAIN
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=c, value=val)
        cell.font = _font(size=9, color=C_GREY_VALUE)
        cell.fill = _fill(bg)
        cell.alignment = _align(h="left")
        cell.border = _border_thin()


def _freeze_and_autofit(ws, freeze_col: int = 4):
    """Freeze top row + first N columns, auto-fit column widths."""
    ws.freeze_panes = ws.cell(row=2, column=freeze_col + 1)
    for col in ws.columns:
        max_w = 10
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ""
                max_w = max(max_w, min(len(val), 60))
            except Exception:
                pass
        ws.column_dimensions[_col_letters(col[0].column)].width = max_w + 2


# ── SUMMARY sheet ─────────────────────────────────────────────────────────────
def _write_summary(wb, model_data: dict):
    ws = wb.create_sheet("SUMMARY", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 28

    pi = model_data.get("project_info", {})
    cats = model_data.get("categories", {})
    src = model_data.get("source", "")
    now = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")

    # Row 1 — big title
    title_cell = ws["A1"]
    title_cell.value = "BIM DATA EXPORT  \u2014  Full Parameter Report"
    title_cell.font = Font(name="Arial", bold=True, size=16, color=C_BLUE_DARK)
    title_cell.fill = _fill(C_BLUE_LIGHT)
    title_cell.alignment = _align(h="left", v="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:D1")

    # Meta rows 2-7
    meta = [
        ("File", src),
        ("Revit Build", pi.get("schema", "IFC2X3")),
        ("Exported", now),
        ("Prepared by", AUTHOR),
        ("Total Categories", len(cats)),
        ("Total Elements", sum(len(v) for v in cats.values())),
    ]
    for r, (label, value) in enumerate(meta, 2):
        la = ws.cell(row=r, column=1, value=label)
        la.font = _font(bold=True, size=9, color=C_GREY_LABEL)
        la.alignment = _align()

        va = ws.cell(row=r, column=2, value=value)
        if label == "Prepared by":
            va.font = _font(bold=True, size=9, color=C_AUTHOR)
        elif label in ("Total Categories", "Total Elements"):
            va.font = _font(bold=False, size=9, color=C_GREY_VALUE)
        else:
            va.font = _font(size=9, color=C_GREY_VALUE)
        va.alignment = _align()

    # Blank row 8
    ws.row_dimensions[8].height = 8

    # Table header row 9
    hdr_cols = ["CATEGORY", "ELEMENTS", "PARAMETERS", "SHEET"]
    for c, col in enumerate(hdr_cols, 1):
        cell = ws.cell(row=9, column=c, value=col)
        cell.font = _font(bold=True, size=10, color=C_HEADER_FG)
        cell.fill = _fill(C_BLUE_DARK)
        cell.alignment = _align(h="center")

    # Data rows 10+
    for r, (cat, rows) in enumerate(cats.items(), 10):
        n_cols = len(CATEGORY_EXTRA_COLS.get(cat, [])) + len(BASE_COLS)
        alt = ((r - 10) % 2 == 0)
        bg = C_ROW_ALT if alt else C_ROW_PLAIN

        for c, val in enumerate([cat, len(rows), n_cols, cat], 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = _fill(bg)
            cell.alignment = _align()
            if c == 2:
                cell.font = _font(bold=True, size=9, color=C_ORANGE)
            elif c in (3, 4):
                cell.font = _font(size=9, color=C_BLUE_DARK)
            else:
                cell.font = _font(size=9, color=C_GREY_VALUE)


# ── Category sheets ────────────────────────────────────────────────────────────
def _write_category_sheet(wb, cat_name: str, rows: list):
    safe_name = cat_name[:31]  # Excel sheet name limit
    ws = wb.create_sheet(safe_name)
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 18

    if not rows:
        ws["A1"] = f"No data for {cat_name}"
        return

    # Collect all columns: BASE_COLS + extra from definition + any extra found in data
    defined_extras = CATEGORY_EXTRA_COLS.get(cat_name, [])
    all_keys = list(BASE_COLS) + [c for c in defined_extras if c not in BASE_COLS]

    # Discover any additional columns from actual data
    data_keys = []
    for row in rows:
        for k in row:
            if k not in all_keys and k not in data_keys:
                data_keys.append(k)
    all_keys += data_keys

    _write_header_row(ws, all_keys, row=1)

    for r_idx, row in enumerate(rows, 2):
        values = [row.get(k, "") for k in all_keys]
        _write_data_row(ws, values, r_idx)

    _freeze_and_autofit(ws, freeze_col=3)


# ── Shared Parameters sheet ─────────────────────────────────────────────────
def _write_shared_params(wb):
    ws = wb.create_sheet("Shared Parameters")
    ws.sheet_view.showGridLines = False

    # Big title (merged, row 1)
    title = ws["A1"]
    title.value = "SharedParameterElement Catalogue  \u2014  QIC (Quality \u00b7 Identity \u00b7 Classification)"
    title.font = Font(name="Arial", bold=True, size=13, color=C_BLUE_DARK)
    title.fill = _fill(C_BLUE_LIGHT)
    title.alignment = _align(h="left", v="center")
    ws.row_dimensions[1].height = 24
    ws.merge_cells("A1:J1")

    # Sub-header row 2
    sub_cols = ["GUID", "Name", "Parameter Type", "Parameter Group", "Group Code",
                "User Modifiable", "Visible", "Discipline", "Description", "Source"]
    _write_header_row(ws, sub_cols, row=2)

    # Data rows
    for r, sp in enumerate(QIC_SHARED_PARAMS, 3):
        alt = ((r - 3) % 2 == 0)
        bg = C_ROW_ALT if alt else C_ROW_PLAIN
        vals = [
            sp["guid"], sp["name"], sp["type"], sp["group"],
            sp["code"], str(sp["mod"]), str(sp["vis"]),
            sp["disc"], sp["desc"], sp["source"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = _fill(bg)
            cell.font = _font(size=9, color=C_GREY_VALUE)
            cell.alignment = _align()
            cell.border = _border_thin()

    _freeze_and_autofit(ws, freeze_col=2)


# ── Family SP Bindings sheet ─────────────────────────────────────────────────
def _write_family_sp_bindings(wb, model_data: dict):
    ws = wb.create_sheet("Family SP Bindings")
    ws.sheet_view.showGridLines = False

    title = ws["A1"]
    title.value = "Family  \u00d7  SharedParameterElement  Bindings  (InstanceBinding + TypeBinding)"
    title.font = Font(name="Arial", bold=True, size=12, color=C_BLUE_DARK)
    title.fill = _fill(C_BLUE_LIGHT)
    title.alignment = _align(h="left", v="center")
    ws.row_dimensions[1].height = 22
    ws.merge_cells("A1:I1")

    sub_cols = ["Family Name", "Category", "IFC Type", "Bound SP Name",
                "GUID", "Parameter Type", "Parameter Group", "Binding Type", "Has Value"]
    _write_header_row(ws, sub_cols, row=2)

    cats = model_data.get("categories", {})
    r_idx = 3
    sp_by_name = {sp["name"]: sp for sp in QIC_SHARED_PARAMS}

    seen_families: set = set()
    for cat, rows in cats.items():
        ifc_type = {v: k for k, v in IFC_CLASS_MAP.items()}.get(cat, "IFCBUILDINGELEMENTPROXY")
        family_names: set = set()
        for row in rows:
            fn = row.get("Object Type", "") or row.get("Name", "").split(":")[0]
            if fn:
                family_names.add(fn)

        for family in sorted(family_names):
            key = f"{cat}:{family}"
            if key in seen_families:
                continue
            seen_families.add(key)
            for sp_name, sp in sp_by_name.items():
                alt = ((r_idx - 3) % 2 == 0)
                bg = C_ROW_ALT if alt else C_ROW_PLAIN
                vals = [
                    family, cat, ifc_type, sp_name,
                    sp["guid"], sp["type"], sp["group"],
                    "InstanceBinding", "No",
                ]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=r_idx, column=c, value=v)
                    cell.fill = _fill(bg)
                    cell.font = _font(size=9, color=C_GREY_VALUE)
                    cell.alignment = _align()
                    cell.border = _border_thin()
                r_idx += 1

    _freeze_and_autofit(ws, freeze_col=2)


# ── Master export function ─────────────────────────────────────────────────────
def export_xlsx(output_path: str, model_data: dict,
                selected_categories: list = None,
                progress_cb=None):
    """
    Export BIM data to Excel in the exact WPH-BUJV reference format.
    progress_cb(percent: int, message: str)
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    cats = model_data.get("categories", {})
    if selected_categories:
        cats = {k: v for k, v in cats.items() if k in selected_categories}

    total_steps = len(cats) + 3  # summary + SP + bindings
    step = 0

    def _progress(msg):
        nonlocal step
        step += 1
        if progress_cb:
            progress_cb(int(100 * step / total_steps), msg)

    # 1. SUMMARY
    _write_summary(wb, {**model_data, "categories": cats})
    _progress("SUMMARY sheet written")

    # 2. Category sheets
    for cat, rows in cats.items():
        _write_category_sheet(wb, cat, rows)
        _progress(f"Sheet: {cat}")

    # 3. Shared Parameters
    _write_shared_params(wb)
    _progress("Shared Parameters sheet written")

    # 4. Family SP Bindings
    _write_family_sp_bindings(wb, {**model_data, "categories": cats})
    _progress("Family SP Bindings sheet written")

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return output_path


def _ensure_pandas():
    """Import pandas, auto-install if missing. Returns pd module."""
    try:
        import pandas as pd
        return pd
    except ImportError:
        import subprocess
        installed = False
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "-q"])
            installed = True
        except Exception:
            pass

        if not installed:
            try:
                subprocess.check_call(["uv", "pip", "install", "pandas"])
                installed = True
            except Exception:
                pass

        if not installed:
            try:
                subprocess.check_call([sys.executable, "-m", "ensurepip", "--default-pip"])
                subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "-q"])
                installed = True
            except Exception:
                pass

        if not installed:
            raise ImportError("Could not install pandas automatically. Please run: uv pip install pandas")

        import pandas as pd
        return pd


def categories_to_dataframes(model_data: dict,
                             pset_filter: set = None,
                             selected_categories: list = None) -> dict:
    """
    Convert model_data categories into a dict of pandas DataFrames.
    One DataFrame per category — columns = all fields present in that category.

    Parameters
    ----------
    model_data          : dict returned by parse_model()
    pset_filter         : set of column-prefix strings to keep, e.g.
                          {"[Pset_WallCommon]", "[Pset_SlabCommon]"}
                          None  → keep ALL columns (no filter)
                          set() → same as None (keep all)
    selected_categories : list of category names to include; None = all

    Returns
    -------
    dict[str, pd.DataFrame]
    """
    pd = _ensure_pandas()
    cats = model_data.get("categories", {})

    if selected_categories:
        cats = {k: v for k, v in cats.items() if k in selected_categories}

    frames: dict = {}
    for cat, rows in cats.items():
        if not rows:
            frames[cat] = pd.DataFrame()
            continue

        df = pd.DataFrame(rows)

        # ── Pset column filter ───────────────────────────────────────────────
        # Columns that look like "[PsetName] PropName" are pset-derived.
        # If pset_filter is supplied, keep only columns whose pset prefix
        # matches something in the filter set.
        if pset_filter:
            def _keep_col(col: str) -> bool:
                m = re.match(r"^\[([^\]]+)\]", col)
                if not m:
                    return True  # base column — always keep
                pset_name = m.group(1)
                return any(pset_name.startswith(p) for p in pset_filter)

            df = df[[c for c in df.columns if _keep_col(c)]]

        # ── Tidy up dtypes ───────────────────────────────────────────────────
        # Element ID should stay as int where possible
        if "Element ID" in df.columns:
            df["Element ID"] = pd.to_numeric(df["Element ID"],
                                             errors="coerce").fillna(0).astype(int)

        # Replace empty strings with pd.NA for cleaner exports
        df = df.replace("", pd.NA)

        frames[cat] = df

    return frames


def export_csv(output_dir: str,
               model_data: dict,
               pset_filter: set = None,
               selected_categories: list = None) -> list:
    """
    Export per-category CSV files using pandas.

    Parameters
    ----------
    output_dir          : folder to write CSVs into
    model_data          : dict from parse_model()
    pset_filter         : optional set of Pset prefix strings to include
                          e.g. {"Pset_WallCommon", "Pset_SlabCommon"}
                          None or set() → export ALL columns
    selected_categories : list of category names; None = all

    Returns
    -------
    list of file paths written
    """
    pd = _ensure_pandas()
    os.makedirs(output_dir, exist_ok=True)

    frames = categories_to_dataframes(model_data, pset_filter, selected_categories)
    paths = []

    for cat, df in frames.items():
        if df.empty:
            continue
        safe = re.sub(r'[\\/:*?"<>|]', "-", cat)
        path = os.path.join(output_dir, f"{safe}.csv")
        df.to_csv(path, index=False, encoding="utf-8-sig")  # utf-8-sig = Excel-friendly BOM
        paths.append(path)

    return paths


def export_csv_single(output_path: str,
                      model_data: dict,
                      pset_filter: set = None,
                      selected_categories: list = None) -> str:
    """
    Export ALL selected categories into ONE combined CSV file.
    Adds a 'Source Category' column so rows stay identifiable.
    """
    pd = _ensure_pandas()
    frames = categories_to_dataframes(model_data, pset_filter, selected_categories)

    combined_frames = []
    for cat, df in frames.items():
        if df.empty:
            continue
        df = df.copy()
        df.insert(0, "Source Category", cat)
        combined_frames.append(df)

    if not combined_frames:
        return ""

    combined = pd.concat(combined_frames, ignore_index=True, sort=False)
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    combined.to_csv(output_path, index=False, encoding="utf-8-sig")
    return output_path


def export_pandas_xlsx(output_path: str,
                       model_data: dict,
                       pset_filter: set = None,
                       selected_categories: list = None) -> str:
    """
    Export using pandas ExcelWriter — one sheet per category.
    Lighter than the full themed openpyxl export; useful for quick data dumps.
    Applies basic header formatting via openpyxl after pandas writes.
    """
    pd = _ensure_pandas()
    frames = categories_to_dataframes(model_data, pset_filter, selected_categories)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for cat, df in frames.items():
            if df.empty:
                continue
            sheet_name = cat[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Apply blue header styling via openpyxl
            ws = writer.sheets[sheet_name]
            hdr_fill = PatternFill("solid", fgColor=C_BLUE_DARK)
            hdr_font = Font(name="Arial", bold=True, size=9, color=C_HEADER_FG)
            hdr_align = Alignment(horizontal="center", vertical="center")
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = hdr_align

            # Alternating row colours
            alt_fill = PatternFill("solid", fgColor=C_ROW_ALT)
            plain_fill = PatternFill("solid", fgColor=C_ROW_PLAIN)
            data_font = Font(name="Arial", size=9)
            for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                bg = alt_fill if r_idx % 2 == 0 else plain_fill
                for cell in row:
                    cell.fill = bg
                    cell.font = data_font

            # Auto column width
            for col in ws.columns:
                max_w = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_w + 2, 60)

    return output_path


def export_ifc(output_path: str, model_data: dict):
    """Export minimal IFC 2x3 STEP Physical File."""
    pi = model_data.get("project_info", {})
    cats = model_data.get("categories", {})
    now = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

    lines = [
        "ISO-10303-21;",
        "HEADER;",
        f"FILE_DESCRIPTION(('BIM Data Export — QIC System v10'),\n  '2;1');",
        f"FILE_NAME('{output_path}','{now}',('{AUTHOR}'),('UCC'),'','','');",
        "FILE_SCHEMA(('IFC2X3'));",
        "ENDSEC;",
        "DATA;",
        f"#1=IFCPROJECT('{str(uuid.uuid4())[:22]}',$,'{pi.get('Project Name', '')}',",
        f"  $,$,$,$,$,$);",
    ]
    eid = 100
    for cat, rows in cats.items():
        for row in rows:
            ifc_cls = row.get("IFC Class", "IFCBUILDINGELEMENTPROXY")
            gid = row.get("IFC GlobalId", str(uuid.uuid4())[:22])
            name = str(row.get("Name", "")).replace("'", "''")
            lines.append(f"#{eid}={ifc_cls}('{gid}',$,'{name}',$,$,$,$);")
            eid += 1

    lines += ["ENDSEC;", "END-ISO-10303-21;"]
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# TKINTER GUI  —  full integrated UI
# ─────────────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
# CSV MODE DIALOG  —  ask user: per-category files or single combined CSV
# ─────────────────────────────────────────────────────────────────────────────
class _CsvModeDialog(tk.Toplevel):
    """Small modal dialog to choose CSV export mode."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("CSV Export Mode")
        self.resizable(False, False)
        self.configure(bg=TK_WHITE)
        self.result = None

        # Centre over parent
        self.transient(parent)
        self.grab_set()

        tk.Label(self, text="Choose CSV export mode",
                 font=("Segoe UI", 10, "bold"),
                 bg=TK_WHITE, fg=TK_BLUE).pack(padx=24, pady=(18, 4))
        tk.Label(self,
                 text="Per-category: one .csv file per category\n"
                      "Combined:     all categories in a single .csv",
                 font=("Segoe UI", 9), bg=TK_WHITE, fg=TK_GREY,
                 justify="left").pack(padx=24, pady=(0, 14))

        btn_row = tk.Frame(self, bg=TK_WHITE)
        btn_row.pack(padx=24, pady=(0, 18))

        for text, val, bg, hv in [
            ("📁  Per-category files", "per_category", TK_BLUE, TK_BLUE_DK),
            ("📄  Single combined CSV", "combined", TK_GREEN, TK_GREEN_HV),
            ("Cancel", None, TK_BORDER, "#bdbdbd"),
        ]:
            b = tk.Button(btn_row, text=text,
                          font=("Segoe UI", 9, "bold"),
                          bg=bg, fg=TK_WHITE if val else TK_GREY,
                          activebackground=hv,
                          activeforeground=TK_WHITE if val else TK_GREY,
                          relief="flat", bd=0, cursor="hand2",
                          padx=12, pady=7,
                          command=lambda v=val: self._pick(v))
            b.bind("<Enter>", lambda e, _b=b, _h=hv: _b.configure(bg=_h))
            b.bind("<Leave>", lambda e, _b=b, _bg=bg: _b.configure(bg=_bg))
            b.pack(side="left", padx=(0, 8))

        self.wait_window(self)

    def _pick(self, val):
        self.result = val
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
# MAIN APPLICATION
# ─────────────────────────────────────────────────────────────────────────────
class App(tk.Tk):
    """Full integrated BIM Export application — v10.0"""

    FONT_MAIN = ("Segoe UI", 9)
    FONT_BOLD = ("Segoe UI", 9, "bold")
    FONT_H1 = ("Segoe UI", 15, "bold")
    FONT_H2 = ("Segoe UI", 10, "bold")
    FONT_SMALL = ("Segoe UI", 8)
    FONT_MONO = ("Consolas", 8)

    def __init__(self):
        super().__init__()
        self.title("BIM DATA EXPORT SYSTEM  v10.0  |  Ahmed Khalaf — UCC BIM Manager")
        self.configure(bg=TK_BG)
        self.geometry("1360x820")
        self.minsize(960, 640)

        # Apply ttk theme overrides
        self._apply_style()

        self.model_data = None
        self.model_path = tk.StringVar(value="")
        self.output_path = tk.StringVar(value="")
        self.status_msg = tk.StringVar(value="Ready  —  open a BIM model to begin")
        self.progress_var = tk.DoubleVar(value=0)
        self._category_vars: dict = {}
        self._pset_vars: dict = {}

        self._build_ui()
        self._update_buttons()
        self._prepopulate_categories()

    # ── ttk style ─────────────────────────────────────────────────────────────
    def _apply_style(self):
        s = ttk.Style(self)
        try:
            s.theme_use("clam")
        except Exception:
            pass

        s.configure("TNotebook", background=TK_BG, borderwidth=0)
        s.configure("TNotebook.Tab", font=self.FONT_BOLD, padding=[14, 5],
                    background=TK_BORDER, foreground=TK_GREY)
        s.map("TNotebook.Tab",
              background=[("selected", TK_WHITE)],
              foreground=[("selected", TK_BLUE)])

        s.configure("Treeview", font=self.FONT_MAIN, rowheight=20,
                    background=TK_WHITE, fieldbackground=TK_WHITE,
                    foreground=TK_GREY, borderwidth=0)
        s.configure("Treeview.Heading", font=self.FONT_BOLD,
                    background=TK_BLUE, foreground=TK_WHITE,
                    relief="flat", borderwidth=0)
        s.map("Treeview.Heading",
              background=[("active", TK_BLUE_DK)])
        s.map("Treeview",
              background=[("selected", TK_BLUE)],
              foreground=[("selected", TK_WHITE)])

        s.configure("Horizontal.TProgressbar",
                    troughcolor=TK_BORDER, background=TK_BLUE,
                    thickness=6, borderwidth=0)
        s.configure("TScrollbar",
                    background=TK_BORDER, troughcolor=TK_BG,
                    arrowcolor=TK_GREY, borderwidth=0)
        s.configure("TCombobox",
                    font=self.FONT_MAIN, fieldbackground=TK_WHITE,
                    background=TK_WHITE, foreground=TK_GREY,
                    selectbackground=TK_BLUE, selectforeground=TK_WHITE)

    # ── UI construction ────────────────────────────────────────────────────────
    def _build_ui(self):
        self._build_header()
        # Main content area
        main = tk.Frame(self, bg=TK_BG)
        main.pack(fill="both", expand=True, padx=0, pady=0)

        pw = tk.PanedWindow(main, orient="horizontal", bg=TK_BG,
                            sashwidth=4, sashrelief="flat",
                            sashpad=2, handlepad=100, handlesize=6)
        pw.pack(fill="both", expand=True, padx=8, pady=8)
        self._build_left_panel(pw)
        self._build_right_panel(pw)
        self._build_footer()

    def _build_header(self):
        hdr = tk.Frame(self, bg=TK_BLUE, height=54)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)

        # Left: logo area
        logo = tk.Frame(hdr, bg=TK_BLUE_DK, width=54, height=54)
        logo.pack(side="left")
        logo.pack_propagate(False)
        tk.Label(logo, text="BIM", font=("Segoe UI", 9, "bold"),
                 bg=TK_BLUE_DK, fg=TK_WHITE).pack(expand=True)

        tk.Label(hdr, text="BIM DATA EXPORT",
                 font=("Segoe UI", 14, "bold"),
                 bg=TK_BLUE, fg=TK_WHITE).pack(side="left", padx=(14, 0))
        tk.Label(hdr, text="  IFC · NWC · RVT → Excel / CSV / IFC",
                 font=("Segoe UI", 9), bg=TK_BLUE, fg="#bbdefb").pack(side="left")

        # Right: Reset + author + version badge
        badge = tk.Frame(hdr, bg="#1557b0", padx=10, pady=4)
        badge.pack(side="right", padx=14, pady=10)
        tk.Label(badge, text="v 10.0", font=("Segoe UI", 8, "bold"),
                 bg="#1557b0", fg=TK_WHITE).pack()

        tk.Label(hdr, text=AUTHOR, font=("Segoe UI", 8),
                 bg=TK_BLUE, fg="#90caf9").pack(side="right", padx=(0, 6))

        # Reset button in header
        self.btn_reset = tk.Button(hdr, text="⟳  Reset All",
                                   command=self._reset_all,
                                   font=("Segoe UI", 8, "bold"),
                                   bg="#d32f2f", fg=TK_WHITE,
                                   activebackground="#b71c1c",
                                   activeforeground=TK_WHITE,
                                   relief="flat", bd=0, cursor="hand2",
                                   padx=10, pady=5)
        self.btn_reset.pack(side="right", padx=(0, 8), pady=12)
        self.btn_reset.bind("<Enter>", lambda e: self.btn_reset.configure(bg="#b71c1c"))
        self.btn_reset.bind("<Leave>", lambda e: self.btn_reset.configure(bg="#d32f2f"))

    # ── Left panel ─────────────────────────────────────────────────────────────
    def _build_left_panel(self, pw):
        left = tk.Frame(pw, bg=TK_PANEL, bd=0,
                        highlightthickness=1, highlightbackground=TK_BORDER)
        pw.add(left, minsize=330, width=355)

        # ── Input section ──
        self._section_label(left, "INPUT MODEL")
        inp = tk.Frame(left, bg=TK_PANEL, padx=10, pady=4)
        inp.pack(fill="x")

        tk.Label(inp, text="Model file  (.ifc  /  .nwc  /  .rvt)",
                 font=self.FONT_SMALL, bg=TK_PANEL, fg=TK_GREY_LT).pack(anchor="w")

        fe = tk.Frame(inp, bg=TK_PANEL)
        fe.pack(fill="x", pady=(2, 6))
        entry = tk.Entry(fe, textvariable=self.model_path,
                         font=self.FONT_MONO, bg="#f1f3f4",
                         relief="flat", bd=0,
                         highlightthickness=1,
                         highlightbackground=TK_BORDER,
                         highlightcolor=TK_BLUE)
        entry.pack(side="left", fill="x", expand=True, ipady=5, padx=(0, 4))
        self._icon_btn(fe, "📂", self._browse_model,
                       tip="Browse for IFC / NWC / RVT").pack(side="right")

        parse_row = tk.Frame(inp, bg=TK_PANEL)
        parse_row.pack(fill="x", pady=(0, 4))
        self.btn_parse = self._pill_btn(
            parse_row, "▶  Parse Model", self._parse_model,
            bg=TK_BLUE, hover=TK_BLUE_DK, fg=TK_WHITE)
        self.btn_parse.pack(side="left", fill="x", expand=True)

        self.btn_fetch_revit = self._pill_btn(
            parse_row, "⚡ Live Revit", self._fetch_live_revit,
            bg=TK_PURPLE, hover=TK_PURPLE_HV, fg=TK_WHITE)
        self.btn_fetch_revit.pack(side="left", fill="x", expand=True, padx=(6, 0))
        if not RUNNING_IN_REVIT:
            self.btn_fetch_revit.configure(state="disabled", bg=TK_BORDER, fg=TK_GREY_LT)

        self.btn_reset_input = self._pill_btn(
            parse_row, "↺", self._reset_all,
            bg="#ef5350", hover="#d32f2f", fg=TK_WHITE, small=True)
        self.btn_reset_input.pack(side="right", padx=(6, 0))

        # ── Output section ──
        self._section_label(left, "OUTPUT")
        out = tk.Frame(left, bg=TK_PANEL, padx=10, pady=4)
        out.pack(fill="x")

        tk.Label(out, text="Export file  (.xlsx)",
                 font=self.FONT_SMALL, bg=TK_PANEL, fg=TK_GREY_LT).pack(anchor="w")

        oe = tk.Frame(out, bg=TK_PANEL)
        oe.pack(fill="x", pady=(2, 8))
        tk.Entry(oe, textvariable=self.output_path,
                 font=self.FONT_MONO, bg="#f1f3f4",
                 relief="flat", bd=0,
                 highlightthickness=1,
                 highlightbackground=TK_BORDER,
                 highlightcolor=TK_BLUE).pack(side="left", fill="x", expand=True,
                                              ipady=5, padx=(0, 4))
        self._icon_btn(oe, "💾", self._browse_output,
                       tip="Choose output path").pack(side="right")

        # Export buttons row
        ef = tk.Frame(out, bg=TK_PANEL)
        ef.pack(fill="x", pady=(0, 2))

        self.btn_xlsx = self._pill_btn(
            ef, "📊 Excel", self._export_xlsx,
            bg=TK_GREEN, hover=TK_GREEN_HV, fg=TK_WHITE)
        self.btn_xlsx.pack(side="left", fill="x", expand=True)

        self.btn_csv = self._pill_btn(
            ef, "📄 CSV", self._export_csv,
            bg=TK_SLATE, hover=TK_SLATE_HV, fg=TK_WHITE)
        self.btn_csv.pack(side="left", fill="x", expand=True, padx=(6, 0))

        self.btn_ifc = self._pill_btn(
            ef, "🏗 IFC", self._export_ifc,
            bg=TK_PURPLE, hover=TK_PURPLE_HV, fg=TK_WHITE)
        self.btn_ifc.pack(side="left", fill="x", expand=True, padx=(6, 0))

        # ── Pset Filter section ──
        self._section_label(left, "PSET FILTER  (columns to include)")
        pf = tk.Frame(left, bg=TK_PANEL, padx=10, pady=6)
        pf.pack(fill="x")

        tk.Label(pf, text="Select property sets  (leave all unchecked = include everything)",
                 font=self.FONT_SMALL, bg=TK_PANEL, fg=TK_GREY_LT,
                 wraplength=300, justify="left").pack(anchor="w", pady=(0, 4))

        # Scrollable pset checkbox area
        pf_scroll = tk.Frame(pf, bg=TK_PANEL,
                             highlightthickness=1, highlightbackground=TK_BORDER)
        pf_scroll.pack(fill="x")
        pf_canvas = tk.Canvas(pf_scroll, bg=TK_PANEL, bd=0,
                              highlightthickness=0, height=130)
        pf_vsb = ttk.Scrollbar(pf_scroll, orient="vertical", command=pf_canvas.yview)
        pf_canvas.configure(yscrollcommand=pf_vsb.set)
        pf_vsb.pack(side="right", fill="y")
        pf_canvas.pack(side="left", fill="both", expand=True)
        self.pset_frame = tk.Frame(pf_canvas, bg=TK_PANEL)
        pf_win = pf_canvas.create_window((0, 0), window=self.pset_frame, anchor="nw")

        def _pf_config(e):
            pf_canvas.configure(scrollregion=pf_canvas.bbox("all"))
            pf_canvas.itemconfig(pf_win, width=pf_canvas.winfo_width())

        self.pset_frame.bind("<Configure>", _pf_config)
        pf_canvas.bind("<Configure>",
                       lambda e: pf_canvas.itemconfig(pf_win, width=e.width))

        # Populate default pset options
        self._pset_vars: dict = {}
        self._populate_pset_checkboxes()

        # Pset filter action buttons
        pfa = tk.Frame(pf, bg=TK_PANEL)
        pfa.pack(fill="x", pady=(4, 0))
        self._pill_btn(pfa, "✓ All", lambda: self._select_all_psets(True),
                       bg=TK_BORDER, hover="#c5cae9",
                       fg=TK_GREY, small=True).pack(side="left")
        self._pill_btn(pfa, "✗ None", lambda: self._select_all_psets(False),
                       bg=TK_BORDER, hover="#c5cae9",
                       fg=TK_GREY, small=True).pack(side="left", padx=(6, 0))
        tk.Label(pfa, text="(none checked = all columns)",
                 font=("Segoe UI", 7), bg=TK_PANEL,
                 fg=TK_GREY_LT).pack(side="left", padx=8)
        catf = tk.Frame(left, bg=TK_PANEL, padx=10, pady=4)
        catf.pack(fill="x")

        # Select all / none buttons
        sa_row = tk.Frame(catf, bg=TK_PANEL)
        sa_row.pack(fill="x", pady=(0, 4))
        self._pill_btn(sa_row, "✓ All", lambda: self._select_all_cats(True),
                       bg=TK_BORDER, hover="#c5cae9", fg=TK_GREY, small=True).pack(side="left")
        self._pill_btn(sa_row, "✗ None", lambda: self._select_all_cats(False),
                       bg=TK_BORDER, hover="#c5cae9", fg=TK_GREY, small=True).pack(side="left", padx=(6, 0))

        # Scrollable checkbox area
        cat_scroll = tk.Frame(catf, bg=TK_PANEL,
                              highlightthickness=1, highlightbackground=TK_BORDER)
        cat_scroll.pack(fill="x")
        canvas = tk.Canvas(cat_scroll, bg=TK_PANEL, bd=0,
                           highlightthickness=0, height=180)
        vsb = ttk.Scrollbar(cat_scroll, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        self.cat_frame = tk.Frame(canvas, bg=TK_PANEL)
        canvas_window = canvas.create_window((0, 0), window=self.cat_frame, anchor="nw")

        def _on_frame_config(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(canvas_window, width=canvas.winfo_width())

        self.cat_frame.bind("<Configure>", _on_frame_config)
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfig(canvas_window, width=e.width))

    # ── Right panel ────────────────────────────────────────────────────────────
    def _build_right_panel(self, pw):
        right = tk.Frame(pw, bg=TK_BG)
        pw.add(right, minsize=520)

        nb = ttk.Notebook(right)
        nb.pack(fill="both", expand=True)

        self.tab_summary = tk.Frame(nb, bg=TK_WHITE)
        nb.add(self.tab_summary, text="   Summary   ")

        self.tab_preview = tk.Frame(nb, bg=TK_WHITE)
        nb.add(self.tab_preview, text="   Data Preview   ")

        self.tab_log = tk.Frame(nb, bg=TK_WHITE)
        nb.add(self.tab_log, text="   Log   ")

        self._build_summary_tab()
        self._build_preview_tab()
        self._build_log_tab()

    def _build_summary_tab(self):
        sf = tk.Frame(self.tab_summary, bg=TK_WHITE, padx=16, pady=12)
        sf.pack(fill="both", expand=True)

        self.lbl_title = tk.Label(sf, text="No model loaded",
                                  font=self.FONT_H1, bg=TK_WHITE, fg=TK_BLUE,
                                  anchor="w")
        self.lbl_title.pack(fill="x")

        self.lbl_meta = tk.Label(sf, text="Open a BIM model (.ifc / .nwc / .rvt) to begin",
                                 font=self.FONT_SMALL, bg=TK_WHITE,
                                 fg=TK_GREY_LT, anchor="w")
        self.lbl_meta.pack(fill="x", pady=(2, 12))

        # Stat cards
        sg = tk.Frame(sf, bg=TK_WHITE)
        sg.pack(fill="x", pady=(0, 16))

        card_defs = [
            ("total_categories", "CATEGORIES", TK_BLUE),
            ("total_elements", "ELEMENTS", TK_ORANGE),
            ("total_params", "PARAM COLS", TK_GREEN),
            ("source_format", "FORMAT", TK_PURPLE),
        ]
        self.stat_labels = {}
        for i, (key, label, color) in enumerate(card_defs):
            card = tk.Frame(sg, bg=color, width=138, height=74)
            card.grid(row=0, column=i, padx=(0, 10))
            card.pack_propagate(False)
            tk.Label(card, text=label,
                     font=("Segoe UI", 7, "bold"),
                     bg=color, fg=TK_WHITE).pack(anchor="w", padx=10, pady=(10, 0))
            lv = tk.Label(card, text="—",
                          font=("Segoe UI", 22, "bold"),
                          bg=color, fg=TK_WHITE, anchor="w")
            lv.pack(fill="x", padx=10)
            self.stat_labels[key] = lv

        # Divider
        tk.Frame(sf, bg=TK_BORDER, height=1).pack(fill="x", pady=(0, 10))

        tk.Label(sf, text="Category Breakdown",
                 font=self.FONT_H2, bg=TK_WHITE, fg=TK_GREY).pack(anchor="w", pady=(0, 6))

        # Summary treeview
        tf = tk.Frame(sf, bg=TK_WHITE,
                      highlightthickness=1, highlightbackground=TK_BORDER)
        tf.pack(fill="both", expand=True)

        cols = ("Category", "Elements", "Parameters", "Sheet")
        self.tree = ttk.Treeview(tf, columns=cols, show="headings", height=10)
        vsb = ttk.Scrollbar(tf, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

        self.tree.heading("Category", text="Category")
        self.tree.heading("Elements", text="Elements")
        self.tree.heading("Parameters", text="Parameters")
        self.tree.heading("Sheet", text="Sheet")
        self.tree.column("Category", width=220, anchor="w")
        self.tree.column("Elements", width=90, anchor="center")
        self.tree.column("Parameters", width=100, anchor="center")
        self.tree.column("Sheet", width=180, anchor="w")
        self.tree.tag_configure("alt", background=TK_ROW_ALT)
        self.tree.tag_configure("plain", background=TK_ROW_PLAIN)

    def _build_preview_tab(self):
        pf = tk.Frame(self.tab_preview, bg=TK_WHITE)
        pf.pack(fill="both", expand=True)

        # Toolbar
        ctrl = tk.Frame(pf, bg=TK_BLUE_LT,
                        highlightthickness=1, highlightbackground=TK_BORDER)
        ctrl.pack(fill="x")

        tk.Label(ctrl, text="Category:",
                 font=self.FONT_BOLD, bg=TK_BLUE_LT, fg=TK_BLUE).pack(
            side="left", padx=(12, 4), pady=7)

        self.preview_cat_var = tk.StringVar()
        self.preview_cat_cb = ttk.Combobox(ctrl, textvariable=self.preview_cat_var,
                                           state="readonly", width=30,
                                           font=self.FONT_MAIN)
        self.preview_cat_cb.pack(side="left", padx=4)
        self.preview_cat_cb.bind("<<ComboboxSelected>>",
                                 lambda e: self._refresh_preview())

        tk.Label(ctrl, text="(first 100 rows shown)",
                 font=self.FONT_SMALL, bg=TK_BLUE_LT, fg=TK_GREY_LT).pack(
            side="left", padx=8)

        # Data treeview
        tf = tk.Frame(pf, bg=TK_WHITE)
        tf.pack(fill="both", expand=True)

        self.preview_tree = ttk.Treeview(tf, show="headings")
        vsb = ttk.Scrollbar(tf, orient="vertical", command=self.preview_tree.yview)
        hsb = ttk.Scrollbar(tf, orient="horizontal", command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=vsb.set,
                                    xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.preview_tree.pack(fill="both", expand=True)
        self.preview_tree.tag_configure("alt", background=TK_ROW_ALT)
        self.preview_tree.tag_configure("plain", background=TK_ROW_PLAIN)

    def _build_log_tab(self):
        lf = tk.Frame(self.tab_log, bg=TK_WHITE)
        lf.pack(fill="both", expand=True, padx=4, pady=4)

        # Toolbar
        tb = tk.Frame(lf, bg=TK_WHITE)
        tb.pack(fill="x", pady=(0, 4))
        self._pill_btn(tb, "🗑 Clear", self._clear_log,
                       bg=TK_BORDER, hover="#cfd8dc", fg=TK_GREY,
                       small=True).pack(side="left")
        tk.Label(tb, text="Application log",
                 font=self.FONT_SMALL, bg=TK_WHITE,
                 fg=TK_GREY_LT).pack(side="left", padx=8)

        self.log_text = tk.Text(lf, font=self.FONT_MONO,
                                bg="#1e1e2e", fg="#cdd6f4",
                                relief="flat", bd=0,
                                wrap="word", state="disabled",
                                insertbackground=TK_WHITE,
                                selectbackground=TK_BLUE)
        sb = ttk.Scrollbar(lf, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True)

        self.log_text.tag_config("ok", foreground="#a6e3a1")
        self.log_text.tag_config("err", foreground="#f38ba8")
        self.log_text.tag_config("warn", foreground="#fab387")
        self.log_text.tag_config("info", foreground="#89dceb")
        self.log_text.tag_config("dim", foreground="#6c7086")

    def _build_footer(self):
        foot = tk.Frame(self, bg=TK_WHITE,
                        highlightthickness=1, highlightbackground=TK_BORDER,
                        height=40)
        foot.pack(fill="x", side="bottom")
        foot.pack_propagate(False)

        self.prog_bar = ttk.Progressbar(foot, variable=self.progress_var,
                                        maximum=100, length=260,
                                        style="Horizontal.TProgressbar",
                                        mode="determinate")
        self.prog_bar.pack(side="left", padx=(12, 0), pady=11)

        tk.Label(foot, textvariable=self.status_msg,
                 font=self.FONT_MAIN, bg=TK_WHITE,
                 fg=TK_GREY).pack(side="left", padx=10)

        # Right: ready indicator
        self.foot_indicator = tk.Label(foot, text="● READY",
                                       font=("Segoe UI", 8, "bold"),
                                       bg=TK_WHITE, fg=TK_GREEN)
        self.foot_indicator.pack(side="right", padx=14)

    # ── Widget helpers ─────────────────────────────────────────────────────────
    def _section_label(self, parent, title: str):
        row = tk.Frame(parent, bg=TK_BLUE_LT)
        row.pack(fill="x", pady=(8, 0))
        tk.Label(row, text=f"  {title}",
                 font=("Segoe UI", 8, "bold"),
                 bg=TK_BLUE_LT, fg=TK_BLUE,
                 height=1).pack(side="left", pady=3)

    def _pill_btn(self, parent, text, cmd,
                  bg=TK_BORDER, hover=None, fg=TK_GREY,
                  full_width=False, small=False) -> tk.Button:
        hover = hover or self._darken(bg)
        padx = 6 if small else 12
        pady = 3 if small else 6
        fnt = self.FONT_SMALL if small else self.FONT_BOLD
        b = tk.Button(parent, text=text, command=cmd,
                      font=fnt, bg=bg, fg=fg,
                      activebackground=hover, activeforeground=fg,
                      relief="flat", bd=0, cursor="hand2",
                      padx=padx, pady=pady)
        b.bind("<Enter>", lambda e, _b=b, _h=hover: _b.configure(bg=_h))
        b.bind("<Leave>", lambda e, _b=b, _bg=bg: _b.configure(bg=_bg))
        return b

    def _icon_btn(self, parent, icon: str, cmd, tip: str = "") -> tk.Button:
        b = tk.Button(parent, text=icon, command=cmd,
                      font=("Segoe UI", 11), bg=TK_BLUE_LT,
                      fg=TK_BLUE, activebackground=TK_BLUE,
                      activeforeground=TK_WHITE,
                      relief="flat", bd=0, cursor="hand2",
                      padx=6, pady=4)
        b.bind("<Enter>", lambda e: b.configure(bg=TK_BLUE, fg=TK_WHITE))
        b.bind("<Leave>", lambda e: b.configure(bg=TK_BLUE_LT, fg=TK_BLUE))
        return b

    @staticmethod
    def _darken(hex_color: str) -> str:
        try:
            h = hex_color.lstrip("#")
            r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
            return f"#{max(r - 24, 0):02x}{max(g - 24, 0):02x}{max(b - 24, 0):02x}"
        except Exception:
            return hex_color

    # ── Logging ────────────────────────────────────────────────────────────────
    def _log(self, msg: str, tag="info"):
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}]  {msg}\n"
        self.log_text.configure(state="normal")
        self.log_text.insert("end", line, tag)
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _clear_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    # ── Actions ────────────────────────────────────────────────────────────────
    def _browse_model(self):
        path = filedialog.askopenfilename(
            title="Open BIM Model",
            filetypes=[
                ("All BIM Files", "*.ifc *.nwc *.rvt"),
                ("IFC Files", "*.ifc"),
                ("Navisworks Cache", "*.nwc"),
                ("Revit Files", "*.rvt"),
                ("All Files", "*.*"),
            ])
        if path:
            self.model_path.set(path)
            base = os.path.splitext(path)[0]
            self.output_path.set(base + "_export.xlsx")
            self._log(f"Selected: {path}", "info")

    def _browse_output(self):
        init = self.output_path.get() or "bim_export.xlsx"
        path = filedialog.asksaveasfilename(
            title="Save Excel Export As",
            initialfile=os.path.basename(init),
            initialdir=os.path.dirname(os.path.abspath(init)),
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("All Files", "*.*")])
        if path:
            self.output_path.set(path)

    def _parse_model(self):
        path = self.model_path.get().strip()
        if not path:
            messagebox.showwarning("No File", "Please select a BIM model file first.")
            return
        if not os.path.exists(path):
            messagebox.showerror("File Not Found",
                                 f"Cannot find file:\n{path}")
            return

        self._set_status("Parsing model…", 2)
        self._set_indicator("⟳ PARSING", TK_ORANGE)
        self._log(f"Parsing: {os.path.basename(path)}", "info")
        self.btn_parse.configure(state="disabled", text="⟳  Parsing…")
        self.update_idletasks()

        import threading
        def _do():
            try:
                md = parse_model(path)
                self.after(0, lambda: self._on_parse_done(md))
            except Exception as e:
                import traceback
                err = traceback.format_exc()
                self.after(0, lambda: self._on_parse_error(str(e), err))

    def _fetch_live_revit(self):
        if not RUNNING_IN_REVIT:
            messagebox.showerror("Not in Revit", "This function is only available when running inside Revit/Dynamo.")
            return

        self._set_status("Fetching live Revit model data…", 2)
        self._set_indicator("⟳ FETCHING", TK_ORANGE)
        self._log("Fetching categories and elements from live Revit document", "info")
        self.btn_fetch_revit.configure(state="disabled", text="⟳ Fetching…")
        
        # Get selected categories from GUI on the main thread
        selected_cats = self._get_selected_categories()
        
        self.update_idletasks()

        import threading
        def _do():
            try:
                # Find active Revit document robustly
                doc = None

                # Method 1: Try Dynamo's DocumentManager
                if 'DocumentManager' in globals():
                    doc = DocumentManager.Instance.CurrentDBDocument
                else:
                    try:
                        from RevitServices.Persistence import DocumentManager
                        doc = DocumentManager.Instance.CurrentDBDocument
                    except Exception:
                        pass

                # Method 2: Try pyRevit / RPS globals
                if doc is None:
                    if '__revit__' in globals():
                        doc = __revit__.ActiveUIDocument.Document
                    elif 'doc' in globals():
                        doc = globals()['doc']
                    else:
                        import __main__
                        if hasattr(__main__, 'doc'):
                            doc = __main__.doc
                        elif hasattr(__main__, '__revit__'):
                            doc = __main__.__revit__.ActiveUIDocument.Document

                if doc is None:
                    raise Exception(
                        "Could not retrieve active Revit Document. Please ensure you are running the script inside Revit (Dynamo, pyRevit, or RPS).")

                # Retrieve category/subcategory data using the script logic
                revit_cats = doc.Settings.Categories
                cats_lines = []
                for cat in revit_cats:
                    if cat is None:
                        continue
                    bic_name = System.Enum.GetName(BuiltInCategory, cat.Id.IntegerValue)
                    bic_str = bic_name if bic_name else "Unknown ({})".format(cat.Id.IntegerValue)

                    cats_lines.append("Category: {}".format(cat.Name))
                    cats_lines.append("  Id: {}".format(cat.Id.IntegerValue))
                    cats_lines.append("  BuiltInCategory: {}".format(bic_str))
                    cats_lines.append("  CategoryType: {}".format(cat.CategoryType))
                    cats_lines.append("  AllowsBoundParameters: {}".format(cat.AllowsBoundParameters))
                    cats_lines.append("  SubCategories:")

                    subcats = cat.SubCategories
                    if subcats is None or subcats.Size == 0:
                        cats_lines.append("    (none)")
                    else:
                        for subcat in subcats:
                            if subcat is None:
                                continue
                            sub_bic_name = System.Enum.GetName(BuiltInCategory, subcat.Id.IntegerValue)
                            sub_bic_str = sub_bic_name if sub_bic_name else "Unknown ({})".format(
                                subcat.Id.IntegerValue)
                            cats_lines.append("    - {}".format(subcat.Name))
                            cats_lines.append("      Id: {}".format(subcat.Id.IntegerValue))
                            cats_lines.append("      BuiltInCategory: {}".format(sub_bic_str))
                            cats_lines.append("      CategoryType: {}".format(subcat.CategoryType))
                    cats_lines.append("")

                # Print to application log
                self.after(0, lambda: self._log(
                    "Revit Categories and Subcategories Information:\n" + "\n".join(cats_lines), "info"))

                p = RVTParser().parse_live(doc, selected_cats)
                md = {
                    "project_info": p.project_info,
                    "categories": p.categories,
                    "shared_param_catalogue": p.shared_param_catalogue,
                    "source": p.source,
                }

                self.after(0, lambda: self._on_parse_done(md))
                self.after(0, lambda: self.btn_fetch_revit.configure(state="normal", text="⚡ Live Revit"))
            except Exception as e:
                import traceback
                err = traceback.format_exc()
                self.after(0, lambda: self._on_parse_error(str(e), err))
                self.after(0, lambda: self.btn_fetch_revit.configure(state="normal", text="⚡ Live Revit"))

        threading.Thread(target=_do, daemon=True).start()

    def _on_parse_done(self, model_data: dict):
        self.model_data = model_data
        cats = model_data.get("categories", {})
        pi = model_data.get("project_info", {})
        src = model_data.get("source", "")
        total_elems = sum(len(v) for v in cats.values())

        self._set_status(
            f"✓  {os.path.basename(src)}  —  {len(cats)} categories  /  {total_elems} elements",
            100)
        self._set_indicator("● LOADED", TK_GREEN)
        self._log(f"Parsed {len(cats)} categories, {total_elems} elements", "ok")
        self._log(f"Project: {pi.get('Project Name', '')}  |  Schema: {pi.get('schema', '')}", "dim")

        # Stat cards
        total_p = sum(len(CATEGORY_EXTRA_COLS.get(c, [])) + len(BASE_COLS) for c in cats)
        self.stat_labels["total_categories"].config(text=str(len(cats)))
        self.stat_labels["total_elements"].config(text=str(total_elems))
        self.stat_labels["total_params"].config(text=str(total_p))
        ext = os.path.splitext(src)[1].upper().lstrip(".") or "—"
        self.stat_labels["source_format"].config(text=ext)

        # Header labels
        self.lbl_title.config(text=pi.get("Project Name", src) or src)
        self.lbl_meta.config(
            text=f"File: {src}   |   Schema: {pi.get('schema', '')}   |   "
                 f"Parsed: {datetime.now().strftime('%Y-%m-%d  %H:%M')}")

        # Summary treeview
        for item in self.tree.get_children():
            self.tree.delete(item)
        for i, (cat, rows) in enumerate(cats.items()):
            n_cols = len(CATEGORY_EXTRA_COLS.get(cat, [])) + len(BASE_COLS)
            tag = "alt" if i % 2 else "plain"
            self.tree.insert("", "end",
                             values=(cat, len(rows), n_cols, cat),
                             tags=(tag,))

        # Category checkboxes
        for w in self.cat_frame.winfo_children():
            w.destroy()
        self._category_vars.clear()
        for i, (cat, rows) in enumerate(cats.items()):
            var = tk.BooleanVar(value=True)
            self._category_vars[cat] = var
            row = tk.Frame(self.cat_frame, bg=TK_ROW_ALT if i % 2 else TK_PANEL)
            row.pack(fill="x")
            tk.Checkbutton(
                row,
                text=f"  {cat}",
                variable=var,
                font=self.FONT_MAIN,
                bg=TK_ROW_ALT if i % 2 else TK_PANEL,
                fg=TK_GREY,
                activebackground=TK_BLUE_LT,
                selectcolor=TK_BLUE_LT,
                relief="flat", bd=0,
                anchor="w",
            ).pack(side="left", fill="x", expand=True, padx=(2, 0))
            tk.Label(row, text=f"{len(rows)}",
                     font=("Segoe UI", 8, "bold"),
                     bg=TK_ROW_ALT if i % 2 else TK_PANEL,
                     fg=TK_ORANGE).pack(side="right", padx=8)

        # Preview combobox
        self.preview_cat_cb["values"] = list(cats.keys())
        if cats:
            self.preview_cat_cb.set(list(cats.keys())[0])
            self._refresh_preview()

        # Refresh pset filter list from discovered psets
        self._refresh_pset_list(model_data)

        self.btn_parse.configure(state="normal", text="▶  Parse Model")
        self._update_buttons()

    def _on_parse_error(self, msg: str, tb: str = ""):
        self._set_status("✗  Parse failed", 0)
        self._set_indicator("● ERROR", TK_ERROR)
        self._log(f"ERROR: {msg}", "err")
        if tb:
            self._log(tb, "err")
        self.btn_parse.configure(state="normal", text="▶  Parse Model")
        messagebox.showerror("Parse Error", f"{msg}\n\nCheck the Log tab for details.")

    def _refresh_preview(self):
        if not self.model_data:
            return
        cat = self.preview_cat_var.get()
        rows = self.model_data["categories"].get(cat, [])

        self.preview_tree.delete(*self.preview_tree.get_children())
        if not rows:
            return

        all_keys = list(rows[0].keys())
        self.preview_tree["columns"] = all_keys
        self.preview_tree["show"] = "headings"

        for k in all_keys:
            w = max(len(k) * 8, 80)
            self.preview_tree.heading(k, text=k)
            self.preview_tree.column(k, width=w, minwidth=60)

        for r_idx, row in enumerate(rows[:100]):
            vals = [str(row.get(k, "")) for k in all_keys]
            tag = "alt" if r_idx % 2 else "plain"
            self.preview_tree.insert("", "end", values=vals, tags=(tag,))

    def _populate_pset_checkboxes(self, discovered: list = None):
        """Build pset filter checkboxes. Uses DEFAULT_PSET_FILTER if no model loaded."""
        for w in self.pset_frame.winfo_children():
            w.destroy()
        self._pset_vars.clear()

        psets = discovered if discovered else sorted(IFCParser.DEFAULT_PSET_FILTER)

        for i, pset in enumerate(psets):
            var = tk.BooleanVar(value=False)  # unchecked = include all
            self._pset_vars[pset] = var
            row = tk.Frame(self.pset_frame,
                           bg=TK_ROW_ALT if i % 2 else TK_PANEL)
            row.pack(fill="x")
            tk.Checkbutton(
                row, text=f"  {pset}",
                variable=var,
                font=self.FONT_SMALL,
                bg=TK_ROW_ALT if i % 2 else TK_PANEL,
                fg=TK_GREY,
                activebackground=TK_BLUE_LT,
                selectcolor=TK_BLUE_LT,
                relief="flat", bd=0, anchor="w",
            ).pack(fill="x", padx=(2, 0))

    def _select_all_psets(self, val: bool):
        for var in self._pset_vars.values():
            var.set(val)

    def _get_pset_filter(self) -> set:
        """
        Return the active pset filter set for use in exports.
        If no boxes are checked → return None (means 'include everything').
        """
        checked = {p for p, var in self._pset_vars.items() if var.get()}
        return checked if checked else None  # None = no filter = all columns

    def _refresh_pset_list(self, model_data: dict):
        """After parsing, update pset checkboxes with psets actually found in model."""
        found_psets = set()
        for rows in model_data.get("categories", {}).values():
            for row in rows:
                for col in row:
                    m = re.match(r"^\[([^\]]+)\]", str(col))
                    if m:
                        found_psets.add(m.group(1))
        if found_psets:
            self._populate_pset_checkboxes(sorted(found_psets))
        else:
            self._populate_pset_checkboxes()

    def _get_selected_categories(self) -> list:
        return [cat for cat, var in self._category_vars.items() if var.get()]

    def _select_all_cats(self, val: bool):
        for var in self._category_vars.values():
            var.set(val)

    # ── Export actions ─────────────────────────────────────────────────────────
    def _export_xlsx(self):
        if not self.model_data:
            messagebox.showwarning("No Data", "Parse a model first.")
            return
        out = self.output_path.get().strip()
        if not out:
            messagebox.showwarning("No Output", "Set an export file path.")
            return
        if not out.lower().endswith(".xlsx"):
            out += ".xlsx"
            self.output_path.set(out)

        selected = self._get_selected_categories()
        if not selected:
            messagebox.showwarning("No Categories", "Select at least one category.")
            return

        self._set_status("Exporting Excel…", 2)
        self._set_indicator("⟳ EXPORTING", TK_ORANGE)
        self._log(f"Exporting Excel → {os.path.basename(out)}", "info")
        self.btn_xlsx.configure(state="disabled", text="⟳ Excel…")

        import threading
        def _do():
            try:
                def cb(pct, msg):
                    self.after(0, lambda: self._set_status(msg, pct))

                export_xlsx(out, self.model_data, selected, progress_cb=cb)
                self.after(0, lambda: self._on_xlsx_done(out))
            except Exception as e:
                import traceback
                self.after(0, lambda: self._on_export_error("Excel", str(e),
                                                            traceback.format_exc()))

        threading.Thread(target=_do, daemon=True).start()

    def _on_xlsx_done(self, path: str):
        sz = os.path.getsize(path) / 1024
        self._set_status(f"✓  Saved: {os.path.basename(path)}  ({sz:.0f} KB)", 100)
        self._set_indicator("● DONE", TK_GREEN)
        self._log(f"✓ Excel saved: {path}  ({sz:.0f} KB)", "ok")
        self.btn_xlsx.configure(state="normal", text="📊 Excel")
        self._update_buttons()
        if messagebox.askyesno("Export Complete",
                               f"File saved:\n{path}\n\nOpen now?"):
            self._open_file(path)

    def _export_csv(self):
        if not self.model_data:
            messagebox.showwarning("No Data", "Parse a model first.")
            return

        # Ask: per-category files or single combined CSV?
        mode = _CsvModeDialog(self).result
        if mode is None:
            return  # user cancelled

        pset_filter = self._get_pset_filter()
        selected = self._get_selected_categories()
        active_filter_str = (
            f"Pset filter: {len(pset_filter)} psets"
            if pset_filter else "Pset filter: ALL columns"
        )

        if mode == "combined":
            base = os.path.splitext(self.output_path.get())[0] or "bim_export"
            path = filedialog.asksaveasfilename(
                title="Save Combined CSV As",
                initialfile=os.path.basename(base) + "_combined.csv",
                defaultextension=".csv",
                filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")])
            if not path:
                return
            try:
                out = export_csv_single(path, self.model_data,
                                        pset_filter=pset_filter,
                                        selected_categories=selected or None)
                sz = os.path.getsize(out) / 1024
                self._log(f"✓ Combined CSV: {out}  ({sz:.0f} KB)  [{active_filter_str}]", "ok")
                messagebox.showinfo("CSV Export",
                                    f"Combined CSV saved:\n{out}\n\n{active_filter_str}")
            except Exception as e:
                self._log(f"CSV error: {e}", "err")
                messagebox.showerror("CSV Error", str(e))

        else:  # per-category
            base = os.path.dirname(self.output_path.get()) or "."
            out_dir = filedialog.askdirectory(
                title="Select CSV Output Folder", initialdir=base)
            if not out_dir:
                return
            try:
                paths = export_csv(out_dir, self.model_data,
                                   pset_filter=pset_filter,
                                   selected_categories=selected or None)
                self._log(
                    f"✓ CSV: {len(paths)} files → {out_dir}  [{active_filter_str}]", "ok")
                messagebox.showinfo(
                    "CSV Export",
                    f"{len(paths)} CSV files saved to:\n{out_dir}\n\n{active_filter_str}")
            except Exception as e:
                self._log(f"CSV error: {e}", "err")
                messagebox.showerror("CSV Error", str(e))

    def _export_ifc(self):
        if not self.model_data:
            messagebox.showwarning("No Data", "Parse a model first.")
            return
        base = os.path.splitext(self.output_path.get())[0] or "export"
        path = filedialog.asksaveasfilename(
            title="Save IFC As",
            initialfile=os.path.basename(base) + ".ifc",
            defaultextension=".ifc",
            filetypes=[("IFC Files", "*.ifc"), ("All Files", "*.*")])
        if not path:
            return
        try:
            export_ifc(path, self.model_data)
            self._log(f"✓ IFC saved: {path}", "ok")
            messagebox.showinfo("IFC Export", f"IFC file saved:\n{path}")
        except Exception as e:
            self._log(f"IFC error: {e}", "err")
            messagebox.showerror("IFC Error", str(e))

    def _on_export_error(self, fmt: str, msg: str, tb: str = ""):
        self._set_status(f"✗  {fmt} export failed", 0)
        self._set_indicator("● ERROR", TK_ERROR)
        self._log(f"ERROR ({fmt}): {msg}", "err")
        if tb:
            self._log(tb, "err")
        self.btn_xlsx.configure(state="normal", text="📊 Excel")
        self._update_buttons()
        messagebox.showerror(f"{fmt} Export Error",
                             f"{msg}\n\nCheck the Log tab for details.")

    @staticmethod
    def _open_file(path: str):
        try:
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                import subprocess;
                subprocess.Popen(["open", path])
            else:
                import subprocess;
                subprocess.Popen(["xdg-open", path])
        except Exception:
            pass

    def _reset_all(self):
        """Reset all UI state and data back to initial empty state."""
        if self.model_data:
            if not messagebox.askyesno(
                    "Reset All",
                    "This will clear all loaded data, paths and logs.\n\nContinue?",
                    icon="warning"):
                return

        # ── Clear data ──
        self.model_data = None

        # ── Clear paths ──
        self.model_path.set("")
        self.output_path.set("")

        # ── Reset stat cards ──
        for lbl in self.stat_labels.values():
            lbl.config(text="—")

        # ── Reset header labels ──
        self.lbl_title.config(text="No model loaded")
        self.lbl_meta.config(
            text="Open a BIM model (.ifc / .nwc / .rvt) to begin")

        # ── Clear summary treeview ──
        for item in self.tree.get_children():
            self.tree.delete(item)

        # ── Repopulate category checkboxes to startup defaults ──
        self._prepopulate_categories()

        # ── Reset pset filter to defaults ──
        self._populate_pset_checkboxes()
        self._select_all_psets(False)  # all unchecked = include everything

        # ── Clear preview treeview ──
        self.preview_tree.delete(*self.preview_tree.get_children())
        self.preview_tree["columns"] = []
        self.preview_cat_cb["values"] = []
        self.preview_cat_var.set("")

        # ── Clear log ──
        self._clear_log()

        # ── Reset progress + status ──
        self.progress_var.set(0)
        self._set_status("Ready  —  open a BIM model to begin", 0)
        self._set_indicator("● READY", TK_GREEN)

        # ── Reset all button states ──
        self.btn_parse.configure(state="normal", text="▶  Parse Model")
        self.btn_xlsx.configure(state="disabled", text="📊 Excel",
                                bg=TK_GREEN)
        self.btn_csv.configure(state="disabled", text="📄 CSV",
                               bg=TK_SLATE)
        self.btn_ifc.configure(state="disabled", text="🏗 IFC",
                               bg=TK_PURPLE)

        self._log("Session reset — ready for new model", "info")

    def _set_status(self, msg: str, pct: float = None):
        self.status_msg.set(msg)
        if pct is not None:
            self.progress_var.set(pct)
        self.update_idletasks()

    def _set_indicator(self, text: str, color: str):
        self.foot_indicator.config(text=text, fg=color)

    def _update_buttons(self):
        state = "normal" if self.model_data else "disabled"
        for btn in (self.btn_csv, self.btn_ifc):
            btn.configure(state=state)
        if self.model_data:
            self.btn_xlsx.configure(state="normal")
    def _prepopulate_categories(self):
        # Prepopulate with all 24 mapped categories in alphabetical order
        categories = sorted(list(CATEGORY_EXTRA_COLS.keys()))
        
        for w in self.cat_frame.winfo_children():
            w.destroy()
        self._category_vars.clear()
        
        for i, cat in enumerate(categories):
            var = tk.BooleanVar(value=True)
            self._category_vars[cat] = var
            row = tk.Frame(self.cat_frame, bg=TK_ROW_ALT if i%2 else TK_PANEL)
            row.pack(fill="x")
            tk.Checkbutton(
                row,
                text=f"  {cat}",
                variable=var,
                font=self.FONT_MAIN,
                bg=TK_ROW_ALT if i%2 else TK_PANEL,
                fg=TK_GREY,
                activebackground=TK_BLUE_LT,
                selectcolor=TK_BLUE_LT,
                relief="flat", bd=0,
                anchor="w",
            ).pack(side="left", fill="x", expand=True, padx=(2, 0))
            tk.Label(row, text="—",
                     font=("Segoe UI", 8),
                     bg=TK_ROW_ALT if i%2 else TK_PANEL,
                     fg=TK_GREY_LT).pack(side="right", padx=8)


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()